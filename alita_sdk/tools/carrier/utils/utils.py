import os
import json
import logging
from dataclasses import dataclass, field

from typing import List, Dict, Any, Optional
import zipfile
import datetime
import io
from enum import Enum

from langchain_core.tools import ToolException
from openpyxl.cell import Cell
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from langchain_core.pydantic_v1 import BaseModel as LangchainBaseModel
from abc import ABC, abstractmethod

import functools
import traceback

# =================================================================================
# MODULE-LEVEL LOGGER
# =================================================================================
logger = logging.getLogger(__name__)


# =================================================================================
# 1. APPLICATION-WIDE CONFIGURATION MODELS (SRP: Storing Configuration)
# =================================================================================
@dataclass(frozen=True)
class GatlingConfig:
    """Configuration constants specific to Gatling report processing."""
    DEFAULT_THINK_TIME: str = "5,0-10,0"
    LOG_LINE_PREFIX_USER: str = "USER"
    LOG_LINE_PREFIX_REQUEST: str = "REQUEST"
    LOG_LINE_PREFIX_GROUP: str = "GROUP"
    LOG_LINE_STATUS_OK: str = "OK"
    LOG_LINE_STATUS_START: str = "START"
    REPORTING_TIMEZONE: str = "US/Eastern"
    REPORT_TYPE: str = "GATLING"
    TIMESTAMP_DIVISOR: int = 1000
    MIN_LOG_PARTS: int = 4
    REQUEST_MIN_PARTS: int = 6
    GROUP_MIN_PARTS: int = 6


"""
Auto-comparison
"""
COMPARISON_SHEET_TITLE = 'comparison'
CONSOLIDATED_REPORT_NAME = 'Comparison_report'
COMPARISON_REPORT = os.path.join('/tmp', CONSOLIDATED_REPORT_NAME)

TESTS_SHEET_TITLE = 'tests'
RESULTS_SHEET_TITLE = 'Test results'
TEXT_FOR_FIRST_DROPDOWN = 'test1 (choose from dropdown)'
TEXT_FOR_SECOND_DROPDOWN = 'test2 (choose from dropdown)'
ADDITIONAL_COLUMN_WIDTH = 5
TRANSACTION_COLUMN_WIDTH = 30
DROPDOWN_COLOR = '00E9F090'
MAIN_HEADER_COLOR = '006ED6EB'
SUB_HEADER_COLOR = '00B8DEE6'
RED_COLOR = 'F7A9A9'
GREEN_COLOR = 'AFF2C9'
YELLOW_COLOR = 'F7F7A9'

LISTS_LIMIT = 30

"""
Auto-comparison
"""
CONSOLIDATED_REPORT_NAME = 'Comparison_report'

COMPARISON_COLUMN_HEADERS = {
    'transaction': "Transaction",
    'req_count': "Req, count",
    'error_count': "KO, count",
    'min': "Min, sec",
    'avg': "Avg, sec",
    'percentile90': "90p, sec",
    'difference': "Difference, 90 pct",
    'max': "Max, sec"
}

COMPARISON_MAIN_HEADERS = {
    'users_count': 'Users',
    'ramp_up': 'Ramp Up, min',
    'duration': 'Duration, min',
    'think_time': 'Think time, sec',
    'start_date': 'Start Date, EST',
    'end_date': 'End Date, EST',
    'throughput': 'Throughput, req/sec',
    'error_rate': 'Error rate, %'
}

COMPARISON_MAIN_HEADERS_LR = {
    'users_count': 'Users',
    'duration': 'Duration, min',
    'start_date': 'Start Date, EST',
    'end_date': 'End Date, EST',
    'throughput': 'Throughput, req/sec',
    'error_rate': 'Error rate, %'
}

LR = "loadrunner"


# =================================================================================
# EXCEL STYLING THEME
# =================================================================================
@dataclass
class ExcelFormattingConfig:
    """Configuration for Excel conditional formatting."""
    # Performance thresholds
    error_rate_warning: float = 5.0
    error_rate_critical: float = 10.0
    response_time_warning: float = 1000.0
    response_time_critical: float = 2000.0
    throughput_warning_threshold: float = 10.0

    colors: Dict[str, str] = field(default_factory=lambda: {
        'excellent': '002BBD4D',  # Green
        'good': 'AFF2C9',  # Light Green
        'warning': 'F7F7A9',  # Yellow
        'critical': 'F7A9A9',  # Red
        'critical_font': '00F90808'  # Red font
    })


@dataclass(frozen=True)
class ExcelStyleTheme:
    """A centralized, immutable theme for Excel report styling - SINGLE SOURCE OF TRUTH"""
    BORDER_THIN = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # === COLOR CONSTANTS (consolidate all scattered colors) ===
    _HEX_GREEN_FILL: str = 'AFF2C9'
    _HEX_YELLOW_FILL: str = 'F7F7A9'
    _HEX_RED_FILL: str = 'F7A9A9'
    HEX_TEAL_HEADER: str = '7FD5D8'
    _HEX_TEAL_SUMMARY: str = 'CDEBEA'
    _HEX_BLUE_FONT: str = '291A75'
    _HEX_GREEN_FONT: str = '2BBD4D'
    _HEX_RED_FONT: str = 'F90808'
    _HEX_GRAY_BORDER: str = '040404'
    _HEX_WHITE_FONT: str = 'FFFFFF'

    # ADD: Consolidate scattered colors from constants
    HEX_DROPDOWN: str = 'E9F090'  # From DROPDOWN_COLOR
    HEX_MAIN_HEADER: str = '6ED6EB'  # From MAIN_HEADER_COLOR
    HEX_SUB_HEADER: str = 'B8DEE6'  # From SUB_HEADER_COLOR

    # === ALL FILLS (no more duplicate PatternFill creation) ===
    # === ALL FILLS (no more duplicate PatternFill creation) ===
    @classmethod
    @property
    def FILL_STATUS_PASSED(cls) -> PatternFill:
        return PatternFill(start_color=cls._HEX_GREEN_FILL, fill_type='solid')

    @classmethod
    @property
    def FILL_STATUS_WARNING(cls) -> PatternFill:
        return PatternFill(start_color=cls._HEX_YELLOW_FILL, fill_type='solid')

    @classmethod
    @property
    def FILL_STATUS_FAILED(cls) -> PatternFill:
        return PatternFill(start_color=cls._HEX_RED_FILL, fill_type='solid')

    @classmethod
    @property
    def FILL_TABLE_HEADER(cls) -> PatternFill:
        return PatternFill(start_color=cls.HEX_TEAL_HEADER, fill_type='solid')

    @classmethod
    @property
    def FILL_SUMMARY_HEADER(cls) -> PatternFill:
        return PatternFill(start_color=cls._HEX_TEAL_SUMMARY, fill_type='solid')

    @classmethod
    @property
    def FILL_DROPDOWN(cls) -> PatternFill:
        return PatternFill(start_color=cls.HEX_DROPDOWN, fill_type='solid')

    @classmethod
    @property
    def FILL_MAIN_HEADER(cls) -> PatternFill:
        return PatternFill(start_color=cls.HEX_MAIN_HEADER, fill_type='solid')

    @classmethod
    @property
    def FILL_SUB_HEADER(cls) -> PatternFill:
        return PatternFill(start_color=cls.HEX_SUB_HEADER, fill_type='solid')

    @classmethod
    @property
    def FONT_HEADER(cls) -> Font:
        return Font(bold=True, color=cls._HEX_WHITE_FONT)

    @classmethod
    @property
    def FONT_SUMMARY_LABEL(cls) -> Font:
        return Font(bold=True, color=cls._HEX_BLUE_FONT)

    @classmethod
    @property
    def FONT_STATUS_PASSED(cls) -> Font:
        return Font(bold=True, color=cls._HEX_GREEN_FONT)

    @classmethod
    @property
    def FONT_STATUS_FAILED(cls) -> Font:
        return Font(bold=True, color=cls._HEX_RED_FONT)

    @classmethod
    @property
    def FONT_HYPERLINK(cls) -> Font:
        return Font(bold=True, underline="single", color=cls._HEX_BLUE_FONT)

    @classmethod
    @property
    def FONT_UI_BOLD(cls) -> Font:
        return Font(bold=True, size=16)

    @classmethod
    @property
    def FONT_UI_HEADER(cls) -> Font:
        return Font(bold=True)

    @classmethod
    @property
    def BORDER_DEFAULT_THIN(cls) -> Border:
        return Border(
            left=Side(style="thin", color=cls._HEX_GRAY_BORDER),
            right=Side(style="thin", color=cls._HEX_GRAY_BORDER),
            top=Side(style="thin", color=cls._HEX_GRAY_BORDER),
            bottom=Side(style="thin", color=cls._HEX_GRAY_BORDER)
        )

    @classmethod
    @property
    def ALIGN_CENTER_ALL(cls) -> Alignment:
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    @classmethod
    @property
    def ALIGN_LEFT_CENTER(cls) -> Alignment:
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    @classmethod
    @property
    def ALIGN_JUSTIFY_TOP(cls) -> Alignment:
        return Alignment(horizontal="justify", vertical="top", wrap_text=True)

    # === FORMATS ===
    FORMAT_PERCENTAGE: str = '0.00%'
    FORMAT_SECONDS: str = '0.000'
    FORMAT_NUMBER: str = '#,##0'
    FORMAT_DECIMAL: str = '#,##0.00'

    @staticmethod
    def get_threshold_rules(rt_threshold: float, er_threshold: float = None):
        """Factory method for creating conditional formatting rules - ELIMINATE DUPLICATION"""
        from openpyxl.formatting.rule import CellIsRule

        rules = list()
        # Response time rules
        rules.append(CellIsRule(
            operator='lessThan',
            formula=[str(rt_threshold)],
            fill=ExcelStyleTheme.FILL_STATUS_PASSED
        ))
        rules.append(CellIsRule(
            operator='greaterThanOrEqual',
            formula=[str(rt_threshold)],
            fill=ExcelStyleTheme.FILL_STATUS_FAILED
        ))

        # Error rate rules if threshold provided
        if er_threshold is not None:
            rules.append(CellIsRule(
                operator='greaterThan',
                formula=[str(er_threshold)],
                fill=ExcelStyleTheme.FILL_STATUS_FAILED
            ))

        return rules

    # === STYLE COMBINATIONS (commonly used together) ===
    @staticmethod
    def apply_header_style(cell):
        """Apply complete header styling to a cell"""
        cell.fill = ExcelStyleTheme.FILL_TABLE_HEADER
        cell.font = ExcelStyleTheme.FONT_HEADER
        cell.border = ExcelStyleTheme.BORDER_DEFAULT_THIN
        cell.alignment = ExcelStyleTheme.ALIGN_CENTER_ALL

    @staticmethod
    def apply_data_style(cell, status: str = None):
        """Apply data cell styling based on status"""
        cell.border = ExcelStyleTheme.BORDER_DEFAULT_THIN
        cell.alignment = ExcelStyleTheme.ALIGN_CENTER_ALL

        if status == 'passed':
            cell.fill = ExcelStyleTheme.FILL_STATUS_PASSED
            cell.font = ExcelStyleTheme.FONT_STATUS_PASSED
        elif status == 'failed':
            cell.fill = ExcelStyleTheme.FILL_STATUS_FAILED
            cell.font = ExcelStyleTheme.FONT_STATUS_FAILED
        elif status == 'warning':
            cell.fill = ExcelStyleTheme.FILL_STATUS_WARNING


class ExcelStyleUtils:
    """Excel styling utilities using REPORT_THEME."""

    @staticmethod
    def get_border(style: str = "thin", color: str = "040404") -> Border:
        """Get consistent border style."""
        border_side = Side(border_style=style, color=color)
        return Border(top=border_side, left=border_side, right=border_side, bottom=border_side)

    @staticmethod
    def get_header_fill(color_key: str = 'header_main') -> PatternFill:
        """Get fill pattern from REPORT_THEME."""
        color_mapping = {
            'header_main': REPORT_THEME.HEX_MAIN_HEADER,
            'header_sub': REPORT_THEME.HEX_SUB_HEADER,
            'header_table': REPORT_THEME.HEX_TEAL_HEADER,
            'dropdown': REPORT_THEME.HEX_DROPDOWN
        }
        color = color_mapping.get(color_key, 'BFBFBF')
        return PatternFill("solid", fgColor=color)

    @staticmethod
    def apply_header_style(cell: Cell, color_key: str = 'header_main'):
        """Apply header styling to a cell."""
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = ExcelStyleUtils.get_header_fill(color_key)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = ExcelStyleUtils.get_border()

    @staticmethod
    def apply_data_style(cell: Cell):
        """Apply data cell styling."""
        cell.font = Font(size=10)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = ExcelStyleUtils.get_border()

    @staticmethod
    def apply_number_style(cell: Cell, decimal_places: int = 2):
        """Apply number formatting to a cell."""
        cell.font = Font(size=10)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = ExcelStyleUtils.get_border()
        cell.number_format = f"#,##0.{'0' * decimal_places}" if decimal_places > 0 else "#,##0"

    @staticmethod
    def apply_percentage_style(cell: Cell, decimal_places: int = 2):
        """Apply percentage formatting to a cell."""
        cell.font = Font(size=10)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = ExcelStyleUtils.get_border()
        cell.number_format = f"0.{'0' * decimal_places}%"

    @staticmethod
    def apply_conditional_color(cell: Cell, value: float, threshold: float,
                                reverse: bool = False):
        """Apply color based on value comparison to threshold.

        Args:
            cell: The cell to style
            value: The numeric value to compare
            threshold: The threshold value
            reverse: If True, values below threshold are bad (red)
        """
        # Define colors based on REPORT_THEME if available
        good_color = "C6EFCE"  # Light green
        bad_color = "FFC7CE"  # Light red
        neutral_color = "FFEB9C"  # Light yellow

        try:
            if reverse:
                # For metrics where lower is better (e.g., error rate, response time)
                if value <= threshold:
                    cell.fill = PatternFill("solid", fgColor=good_color)
                elif value <= threshold * 1.2:  # Within 20% of threshold
                    cell.fill = PatternFill("solid", fgColor=neutral_color)
                else:
                    cell.fill = PatternFill("solid", fgColor=bad_color)
            else:
                # For metrics where higher is better (e.g., throughput)
                if value >= threshold:
                    cell.fill = PatternFill("solid", fgColor=good_color)
                elif value >= threshold * 0.8:  # Within 20% of threshold
                    cell.fill = PatternFill("solid", fgColor=neutral_color)
                else:
                    cell.fill = PatternFill("solid", fgColor=bad_color)
        except (ValueError, TypeError):
            # If comparison fails, don't apply color
            pass

    @staticmethod
    def auto_adjust_column_width(worksheet, min_width: int = 10, max_width: int = 50):
        """Auto-adjust column widths based on content."""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max(max_length + 2, min_width), max_width)
            worksheet.column_dimensions[column_letter].width = adjusted_width


# Update the global instance
REPORT_THEME = ExcelStyleTheme()

# Global instances for easy, consistent import.
GATLING_CONFIG = GatlingConfig()

def aggregate_errors(test_errors: list) -> dict:
    """Aggregates error counts from a list of error dictionaries."""
    aggregated_errors = {}
    for error_dict in test_errors:
        for err, details in error_dict.items():
            if err not in aggregated_errors:
                aggregated_errors[err] = details.copy()
            else:
                count = aggregated_errors[err].get('Error count', 0)
                aggregated_errors[err]['Error count'] = int(count) + int(details.get('Error count', 0))
    return aggregated_errors


def exclude_from_dict(data: dict, keys_to_exclude_str: str) -> dict:
    """Removes keys from a dictionary if they contain any of the substrings."""
    if not keys_to_exclude_str:
        return data
    exclude_substrings = list(filter(None, keys_to_exclude_str.split(',')))

    filtered_data = data.copy()
    for sub in exclude_substrings:
        filtered_data = {k: v for k, v in filtered_data.items() if sub not in k}
    return filtered_data


def parse_config_from_string(config_str: str) -> dict:
    """Parses configuration from a JSON string or a key:value newline-separated string."""
    try:
        # First, attempt to parse as a valid JSON object.
        return json.loads(config_str)
    except (json.JSONDecodeError, TypeError):
        # If it fails, fall back to parsing as key: value pairs.
        config = {}
        if isinstance(config_str, str):
            for line in config_str.splitlines():
                if ':' in line:
                    key, value = line.split(':', 1)
                    config[key.strip()] = value.strip()
        return config


def get_latest_log_file(root_dir: str, log_file_name: str) -> str:
    """Finds the most recently modified sub-folder in a directory and returns the path to a log file within it."""
    if not os.path.isdir(root_dir):
        logger.error(f"Root directory for logs not found: {root_dir}")
        raise FileNotFoundError(f"Directory not found: {root_dir}")

    try:
        # Get all subdirectories and sort by modification time.
        subfolders = [os.path.join(root_dir, d) for d in os.listdir(root_dir) if
                      os.path.isdir(os.path.join(root_dir, d))]
        if not subfolders:
            raise FileNotFoundError(f"No subdirectories found in {root_dir}")

        latest_folder = max(subfolders, key=os.path.getmtime)
        simulation_log_file = os.path.join(latest_folder, log_file_name)

        if not os.path.isfile(simulation_log_file):
            logger.error(f"Log file not found in latest directory: {simulation_log_file}")
            raise FileNotFoundError(f"File not found: {simulation_log_file}")

        logger.info(f"Found latest log file: {simulation_log_file}")
        return simulation_log_file

    except Exception as e:
        logger.error(f"Error while searching for log file in '{root_dir}': {e}", exc_info=True)
        raise


@dataclass(frozen=True)
class AnalysisRulesConfig:
    """
    Configuration for the business rules used in performance analysis.
    This centralizes all "magic numbers" for easy tuning.
    """
    # Response Time Severity Multipliers (based on the primary threshold)
    RT_CRITICAL_MULTIPLIER: float = 3.0
    RT_HIGH_MULTIPLIER: float = 2.0

    # Error Rate Severity Multiplier
    ER_HIGH_MULTIPLIER: float = 2.0

    # Throughput Severity Threshold (as a percentage below target)
    TP_HIGH_SEVERITY_DEFICIT_PCT: float = 50.0

    # System Reliability Thresholds (as a percentage success rate)
    RELIABILITY_STANDARD_SUCCESS_RATE: float = 99.0
    RELIABILITY_HIGH_PRIORITY_THRESHOLD: float = 95.0

    # Imbalance Detection
    IMBALANCE_MIN_TRANSACTIONS: int = 5
    IMBALANCE_FACTOR: float = 5.0  # How many times slower the slowest are compared to the fastest

    # Maximum number of recommendations to return
    MAX_RECOMMENDATIONS: int = 5


# Add a global instance for easy import
ANALYSIS_CONFIG = AnalysisRulesConfig()


class CarrierArtifactUploader:
    """
    A dedicated, reusable utility for handling file I/O and uploading artifacts
    to Carrier. It encapsulates the legacy requirement of zipping files before upload.
    """

    def __init__(self, api_wrapper: Any):
        if not api_wrapper:
            raise ToolException("CarrierArtifactUploader requires a valid api_wrapper")
        self.api = api_wrapper
        self.logger = logging.getLogger(__name__)

    def upload(self, bucket_name: str, filename: str, file_bytes: bytes) -> bool:
        """
        Direct upload method that matches the expected interface.
        This is what your loader is trying to call.
        """
        return self.upload_from_bytes(file_bytes, bucket_name, filename)

    def upload_from_bytes(self, file_bytes: bytes, bucket_name: str, filename: str) -> bool:
        """Upload file from bytes directly to Carrier."""
        try:
            # Use the API wrapper's method directly
            self.api.upload_report_from_bytes(file_bytes, bucket_name, filename)
            self.logger.info(f"Successfully uploaded {filename} to {bucket_name}")
            return True
        except Exception as e:
            self.logger.error(f"Failed to upload {filename}: {e}")
            return False

    def upload_leg(self, file_bytes: bytes, bucket_name: str, remote_filename: str) -> bool:
        """
        Handles the complete upload process, including zipping, saving to a
        temporary file, and cleanup, based on the legacy requirement.

        Args:
            file_bytes: The raw bytes of the file to upload (e.g., the Excel report).
            bucket_name: The target bucket in Carrier.
            remote_filename: The desired filename on the remote server (e.g., "report.xlsx").

        Returns:
            True if the upload was successful, False otherwise.

        Raises:
            ToolException: If upload fails for any reason.
        """
        if not file_bytes:
            raise ToolException("File bytes cannot be empty")
        if not bucket_name:
            raise ToolException("Bucket name is required")
        if not remote_filename:
            raise ToolException("Remote filename is required")

        temp_zip_path = f"/tmp/{os.path.splitext(remote_filename)[0]}.zip"
        self.logger.info(
            f"Preparing to upload '{remote_filename}' by creating a temporary zip archive at '{temp_zip_path}'.")

        # Step 1: Create a zip archive in memory
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
            # Add the Excel file bytes to the zip archive with its desired name
            zipf.writestr(remote_filename, file_bytes)
        zip_bytes = zip_buffer.getvalue()

        # Step 2: Write the zip archive bytes to a temporary file
        with open(temp_zip_path, 'wb') as temp_file:
            temp_file.write(zip_bytes)

        self.logger.debug(f"Temporary zip file '{temp_zip_path}' created successfully.")

        # Step 3: Upload the temporary zip file using the provided API wrapper method
        success = self.api.upload_file(bucket_name, temp_zip_path)

        if not success:
            raise ToolException(f"Failed to upload '{temp_zip_path}' to bucket '{bucket_name}'")

        self.logger.info(f"Successfully uploaded '{temp_zip_path}' to bucket '{bucket_name}'.")
        return True

    def generate_download_url(self, bucket_name: str, filename: str) -> str:
        """Generate download URL for the uploaded file."""
        try:
            # Get project ID and base URL from API wrapper
            if hasattr(self.api, '_client') and self.api._client:
                project_id = self.api.project_id
                base_url = self.api.url.rstrip('/')
            elif hasattr(self.api, 'carrier_client'):
                project_id = self.api.carrier_client.credentials.project_id
                base_url = self.api.carrier_client.credentials.url.rstrip('/')
            else:
                # Fallback for different API wrapper structure
                project_id = getattr(self.api, 'project_id', 'default')
                base_url = getattr(self.api, 'url', getattr(self.api, 'base_url', ''))

            # Generate the download URL matching Carrier's API structure
            download_url = f"{base_url}/api/v1/artifacts/artifact/default/{project_id}/{bucket_name}/{filename}?integration_id=1&is_local=False"

            logger.info(f"Generated download URL: {download_url}")
            return download_url

        except Exception as e:
            logger.warning(f"Could not generate download link: {e}")
            return f"Upload successful. Access file '{filename}' in bucket '{bucket_name}'."

class ReportInsights(LangchainBaseModel):
    """Performance report analysis insights"""
    report_type: str
    key_metrics: List[str]
    issues_found: List[str]
    recommendations: List[str]
    comparison_insights: Optional[List[str]] = None


class ReportAnalysis(LangchainBaseModel):
    """Complete report analysis structure"""
    insights: ReportInsights
    confidence_score: float
    analysis_timestamp: str


# ==============================================================================
# INTERFACE DEFINITIONS - Interface Segregation Principle
# ==============================================================================

class IReportProcessor(ABC):
    """Interface for report processing operations"""

    @abstractmethod
    def process_report(self, report_data: Dict[str, Any]) -> Dict[str, Any]:
        pass

    @abstractmethod
    def generate_insights(self, report_data: Dict[str, Any]) -> ReportAnalysis:
        pass


# ==============================================================================
# CONSTANTS & ENUMS - DRY Principle
# ==============================================================================

class FileType(Enum):
    """File type enumeration"""
    HTML = ".html"
    JSON = ".json"
    EXCEL = ".xlsx"
    ZIP = ".zip"


def tool_logger(func):
    """
    A decorator that provides extended, structured logging for tool execution.
    It logs entry, exit, arguments, return values, and exceptions.
    """

    @functools.wraps(func)
    def wrapper(self, *args, **kwargs):
        tool_name = self.name
        # Combine args and kwargs for a complete picture of the input
        all_args = {**kwargs}
        # Handle positional arguments if any (rare for these tools)
        arg_names = func.__code__.co_varnames[1:func.__code__.co_argcount]  # [1:] to skip 'self'
        for i, arg in enumerate(args):
            all_args[arg_names[i]] = arg

        logger.info(f"[{tool_name}] ---> Entering tool execution.")
        logger.debug(f"[{tool_name}] Arguments: {all_args}")

        try:
            result = func(self, *args, **kwargs)
            logger.debug(f"[{tool_name}] Raw result: {str(result)[:250]}...")  # Log a preview of the result
            logger.info(f"[{tool_name}] <--- Tool execution successful.")
            return result
        except Exception as e:
            # Use ToolException for controlled errors, log others more severely
            if isinstance(e, ToolException):
                logger.warning(f"[{tool_name}] <--- Tool execution failed with ToolException: {e}")
            else:
                # Log the full stack trace for unexpected errors
                stack_trace = traceback.format_exc()
                logger.error(f"[{tool_name}] <--- Tool execution failed with unhandled exception: {e}\n{stack_trace}")
            # Re-raise the exception so the framework can handle it
            raise

    return wrapper


class DateTimeUtils:
    """Centralized date/time parsing utilities."""

    SUPPORTED_FORMATS = [
        '%Y-%m-%d %H:%M:%S',
        '%Y-%m-%d %H:%M:%S.%f',
        '%Y-%m-%dT%H:%M:%S',
        '%Y-%m-%dT%H:%M:%S.%fZ',
        '%Y-%m-%d',
        '%m/%d/%Y %H:%M:%S',
        '%d/%m/%Y %H:%M:%S',
        '%Y%m%d_%H%M%S'
    ]

    @classmethod
    def parse_datetime(cls, date_str: str, default: Optional[datetime] = None) -> datetime:
        """Parse datetime string with multiple format support."""
        from datetime import datetime
        if not date_str:
            return default or datetime.now()

        if isinstance(date_str, datetime):
            return date_str

        for fmt in cls.SUPPORTED_FORMATS:
            try:
                return datetime.strptime(str(date_str).strip(), fmt)
            except (ValueError, TypeError):
                continue

        return default or datetime.now()

    @classmethod
    def get_current_timestamp(cls) -> str:
        """Get current timestamp as formatted string."""
        from datetime import datetime
        return datetime.now().strftime('%Y%m%d_%H%M%S')

class FileNameGenerator:
    """Consistent file naming across the framework."""

    @staticmethod
    def generate_report_name(test_name: str, file_type: FileType, timestamp: Optional[datetime] = None) -> str:
        """Generate consistent file names."""
        from datetime import datetime
        ts = timestamp or datetime.now()
        timestamp_str = ts.strftime('%Y%m%d_%H%M')
        base_name = f"consolidated_comparison_{test_name}_{timestamp_str}"
        return f"{base_name}.{file_type.value}"

    @staticmethod
    def generate_comparison_filename(test_name: str, file_type: FileType) -> str:
        """Generate filename for comparison report."""
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        extension = "xlsx" if file_type == FileType.EXCEL else "png"
        return f"{test_name}_comparison_{timestamp}.{extension}"
