import io
import os
import logging
from typing import Dict, Any, Optional
from datetime import datetime

from ..etl_pipeline import BaseLoader
from ..reporting.core.data_models import PerformanceReport
from ..reporting.backend_excel_reporter import ExcelReporter
from langchain_core.tools import ToolException
from alita_sdk.tools.carrier.utils.utils import CarrierArtifactUploader


class CarrierExcelLoader(BaseLoader):
    """
    🎯 Excel loader that generates reports and provides download links.
    Uses existing API wrapper methods following DRY principles.
    """

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.logger = logging.getLogger(__name__)

    def load(self, transformed_data: PerformanceReport, context: Dict[str, Any]) -> Dict[str, Any]:
        """
        💾 Generates Excel report and uploads as ZIP using legacy pattern.
        """
        self.logger.info("Initiating Loader phase...")

        try:
            # Validate context
            api_wrapper = context.get("api_wrapper")
            report_metadata = context.get("report_metadata", {})

            if not api_wrapper:
                raise ToolException("API wrapper missing from context.")

            # Step 1: Generate Excel report in memory
            self.logger.info("📊 Generating Excel report...")
            excel_reporter = ExcelReporter()
            report_workbook = excel_reporter.generate_workbook(transformed_data)

            # Convert to bytes
            excel_buffer = io.BytesIO()
            report_workbook.save(excel_buffer)
            excel_bytes = excel_buffer.getvalue()
            excel_size = len(excel_bytes)

            self.logger.info(f"✅ Successfully rendered Excel report ({excel_size} bytes).")

            # Step 2: Prepare upload details
            bucket_name, file_name = self._get_upload_details(report_metadata)

            # Step 3: Use legacy ZIP upload method
            uploader = CarrierArtifactUploader(api_wrapper)

            self.logger.info(f"📤 Uploading '{file_name}' as ZIP to bucket '{bucket_name}'...")

            upload_success = uploader.upload_leg(excel_bytes, bucket_name, file_name)

            if not upload_success:
                raise ToolException("Report was generated but upload failed.")

            # Step 4: ✅: Generate ZIP download link
            zip_file_name = f"{os.path.splitext(file_name)[0]}.zip"  # Change .xlsx to .zip
            download_url = self._generate_download_link(api_wrapper, bucket_name, zip_file_name)

            # Step 5: Return comprehensive result
            self.logger.info("✅ Loading phase completed successfully.")

            return {
                "status": "success",
                "report_url": getattr(transformed_data, 'carrier_report_url', 'Unknown'),
                "message": "Excel report generated and uploaded as ZIP successfully",
                "file_name": zip_file_name,
                "excel_file_name": file_name,
                "bucket_name": bucket_name,
                "file_size_bytes": excel_size,
                "download_url": download_url,
                "upload_success": upload_success,
                "metadata": {
                    "report_id": report_metadata.get("id"),
                    "report_name": report_metadata.get("name"),
                    "build_status": getattr(transformed_data, 'build_status', 'Unknown'),
                    "timestamp": datetime.now().isoformat(),
                    "loader_type": "CarrierExcelLoader",
                    "archive_type": "ZIP"
                }
            }

        except Exception as e:
            self.logger.error(f"💥 Loading phase failed: {e}", exc_info=True)
            raise ToolException(f"Excel loading failed: {str(e)}")

    def _get_upload_details(self, metadata: dict) -> tuple[str, str]:
        """
        📋 Extract upload details from metadata using existing patterns.
        """
        # Use same bucket naming as existing code
        bucket_name = metadata.get("name", "default_bucket").replace("_", "").replace(" ", "").lower()

        # Generate filename using existing pattern
        build_id = metadata.get("build_id")
        if not build_id:
            # Fallback to report ID if no build_id
            report_id = metadata.get("id", "unknown")
            build_id = f"{report_id}"

        file_name = f"reports_test_results_{build_id}_excel_report.xlsx"

        self.logger.debug(f"📁 Upload details: bucket='{bucket_name}', file='{file_name}'")
        return bucket_name, file_name

    def _generate_download_link(self, api_wrapper, bucket_name: str, file_name: str) -> str:
        """
        Generates the artifact download URL using the same logic as CarrierExcelLoader.
        """
        try:
            # This logic is copied from the working CarrierExcelLoader
            if hasattr(api_wrapper, 'carrier_client'):
                project_id = api_wrapper.carrier_client.credentials.project_id
                base_url = api_wrapper.carrier_client.credentials.url.rstrip('/')
            else:
                project_id = api_wrapper._client.credentials.project_id
                base_url = api_wrapper._client.credentials.url.rstrip('/')

            download_url = f"{base_url}/api/v1/artifacts/artifact/default/{project_id}/{bucket_name}/{file_name}?integration_id=1&is_local=False"
            self.logger.info(f"Successfully generated download link: {download_url}")
            return download_url
        except Exception as e:
            self.logger.warning(f"Could not generate download link: {e}. Providing manual access info.")
            return f"Upload successful. Access file '{file_name}' in bucket '{bucket_name}'."


class CarrierPPTXLoader(BaseLoader):
    """
    📊 PowerPoint loader - placeholder for future implementation.
    """

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.logger = logging.getLogger(__name__)

    def load(self, transformed_data: Any, context: Dict[str, Any]) -> Dict[str, Any]:
        self.logger.warning("📊 CarrierPPTXLoader is not yet implemented.")
        raise NotImplementedError("PPTX loading not implemented yet")


class CarrierDocxLoader(BaseLoader):
    """
    📝 Word document loader - placeholder for future implementation.
    """

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.logger = logging.getLogger(__name__)

    def load(self, transformed_data: Any, context: Dict[str, Any]) -> Dict[str, Any]:
        self.logger.warning("📝 CarrierDocxLoader is not yet implemented.")
        raise NotImplementedError("DOCX loading not implemented yet")


import zipfile
from openpyxl.formatting.formatting import ConditionalFormattingList

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Side, Alignment, Border
from openpyxl.formatting.rule import CellIsRule
from copy import copy

class ComparisonExcelLoader(BaseLoader):
    """
    A stateful loader that manages a persistent, consolidated comparison report.
    This version includes the full logic to build the interactive dashboard.
    """

    def __init__(self, history_limit: int = 20):
        super().__init__()
        self.logger = logging.getLogger(__name__)
        self.history_limit = history_limit
        self.reporter = ExcelReporter()

    def load(self, transformed_data: Dict[str, Any], context: Dict[str, Any]) -> Dict[str, Any]:
        self.logger.info("Starting stateful comparison loading process...")
        api_wrapper = context["api_wrapper"]
        test_name = context.get("test_name", "default_test")
        zip_report_name = f"consolidated_comparison_{test_name}.zip"
        bucket_name = context.get("comparison_bucket", "performance-comparisons")
        try:
            excel_workbook, _ = self._get_or_create_consolidated_workbook(api_wrapper, bucket_name,
                                                                          f"consolidated_comparison_{test_name}.xlsx")
            for report in transformed_data.get("all_reports", []):
                sheet_name = self._add_report_as_hidden_sheet(excel_workbook, report)
                self._update_tests_index(excel_workbook, sheet_name, report)
            self._prune_old_reports(excel_workbook)
            self._rebuild_comparison_dashboard(excel_workbook)
            ai_analysis_md = self._create_ai_analysis_markdown(transformed_data.get("comparison_analysis"))
            download_url = self._create_and_upload_zip_archive(api_wrapper, excel_workbook, ai_analysis_md, bucket_name,
                                                               zip_report_name)
            self.logger.info("Consolidated ZIP archive successfully created and uploaded.")
            return {"status": "success",
                    "message": "Consolidated comparison report and AI analysis ZIP archive created successfully.",
                    "download_url": download_url}
        except Exception as e:
            self.logger.error(f"Stateful comparison loading failed: {e}", exc_info=True)
            return {"status": "error", "error": str(e)}

    def _fill_header(self, ws, row, col, value, color="BFBFBF", merge_cols=0):
        """Helper to fill and style a header cell."""
        cell = ws.cell(row=row, column=col)
        cell.value = value
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill("solid", fgColor=color)
        border_style = Side(border_style="thin", color="040404")
        cell.border = Border(top=border_style, left=border_style, right=border_style, bottom=border_style)
        if merge_cols > 0:
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + merge_cols)

    def _set_cell_formula(self, ws, row, col, formula):
        """Helper to set a formula and style a data cell."""
        cell = ws.cell(row=row, column=col)
        cell.value = formula
        border_style = Side(border_style="thin", color="040404")
        cell.border = Border(top=border_style, left=border_style, right=border_style, bottom=border_style)

    def _apply_conditional_formatting(self, ws, cell_range, low_threshold, high_threshold):
        """Helper to apply standard conditional formatting."""
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type='solid')
        yellow_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type='solid')
        red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type='solid')
        ws.conditional_formatting.add(cell_range,
                                      CellIsRule(operator='between', formula=[low_threshold, high_threshold],
                                                 fill=yellow_fill))
        ws.conditional_formatting.add(cell_range,
                                      CellIsRule(operator='lessThanOrEqual', formula=[low_threshold], fill=green_fill))
        ws.conditional_formatting.add(cell_range,
                                      CellIsRule(operator='greaterThan', formula=[high_threshold], fill=red_fill))

    def _rebuild_comparison_dashboard(self, wb: Workbook):
        """Re-creates the dynamic comparison sheet using formulas, porting all legacy logic."""
        self.logger.info("Rebuilding full dynamic comparison dashboard...")
        comparison_ws = wb["Comparison"]
        tests_ws = wb["tests"]

        comparison_ws.delete_rows(1, comparison_ws.max_row + 1)
        comparison_ws.conditional_formatting = ConditionalFormattingList()

        if tests_ws.max_row < 2:
            comparison_ws[
                "A1"] = "Not enough reports in history to build a comparison. Please run at least one more test."
            self.logger.warning("Not enough reports to build a comparison dashboard. Need at least 2.")
            return

        # 1. Setup Dropdowns
        data_val = DataValidation(type="list", formula1='=tests!A:A')
        comparison_ws.add_data_validation(data_val)
        data_val.add(comparison_ws["H1"])
        data_val.add(comparison_ws["H2"])
        comparison_ws["H1"] = tests_ws.cell(row=tests_ws.max_row - 1, column=1).value
        comparison_ws["H2"] = tests_ws.cell(row=tests_ws.max_row, column=1).value
        comparison_ws.cell(row=1, column=8).fill = PatternFill("solid", fgColor="DDEBF7")
        comparison_ws.cell(row=2, column=8).fill = PatternFill("solid", fgColor="DDEBF7")

        # 2. Build Header Section
        headers = ["Users", "Ramp up, sec", "Duration, min", "Think time, sec", "Start Date, BST", "End Date, BST",
                   "Throughput, req/sec", "Error rate, %"]
        for i, header in enumerate(headers, 1):
            comparison_ws.cell(row=i + 3, column=1).value = header
            self._set_cell_formula(comparison_ws, row=i + 3, col=2,
                                   formula=f'=IFERROR(VLOOKUP("{header}", INDIRECT("\'"&$H$1&"\'!A:B"), 2, FALSE), "N/A")')
            self._set_cell_formula(comparison_ws, row=i + 3, col=4,
                                   formula=f'=IFERROR(VLOOKUP("{header}", INDIRECT("\'"&$H$2&"\'!A:B"), 2, FALSE), "N/A")')

        # 3. Build Main Table Headers
        table_start_row = len(headers) + 6
        header_row = table_start_row - 1
        self._fill_header(comparison_ws, header_row, 1, "Transaction")
        self._fill_header(comparison_ws, header_row, 2, "Req. count", merge_cols=1)
        self._fill_header(comparison_ws, header_row, 5, "KO, count", merge_cols=1)
        self._fill_header(comparison_ws, header_row, 8, "Min, sec", merge_cols=1)
        self._fill_header(comparison_ws, header_row, 11, "Avg, sec", merge_cols=1)
        self._fill_header(comparison_ws, header_row, 14, "90p, sec", merge_cols=1)
        self._fill_header(comparison_ws, header_row, 17, "Difference, 90 pct", merge_cols=1)
        self._fill_header(comparison_ws, header_row, 20, "Max, sec", merge_cols=1)

        for col in [2, 5, 8, 11, 14, 20]: self._fill_header(comparison_ws, table_start_row, col, "test1", "D9EAD3")
        for col in [3, 6, 9, 12, 15, 21]: self._fill_header(comparison_ws, table_start_row, col, "test2", "D9EAD3")
        self._fill_header(comparison_ws, table_start_row, 17, "Diff, sec", "D9EAD3")
        self._fill_header(comparison_ws, table_start_row, 18, "Diff, %", "D9EAD3")

        # 4. Populate Transaction Names
        all_transactions = set()
        for i in range(1, tests_ws.max_row + 1):
            sheet_name = tests_ws.cell(row=i, column=1).value
            if sheet_name and sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                for row in range(13, sheet.max_row + 1):
                    if sheet.cell(row=row, column=1).value:
                        all_transactions.add(sheet.cell(row=row, column=1).value)

        current_row = table_start_row + 1
        for tx_name in sorted(list(all_transactions)):
            comparison_ws.cell(row=current_row, column=1).value = tx_name
            current_row += 1

        # 5. Populate Table with Formulas
        for r in range(table_start_row + 1, comparison_ws.max_row + 1):
            tx_name_cell = f"$A{r}"
            # Req Count, KO Count
            for i, col_idx in enumerate([2, 3], start=2):
                self._set_cell_formula(comparison_ws, r, i,
                                       f'=IFERROR(VLOOKUP({tx_name_cell},INDIRECT("\'"&$H$1&"\'!A:J"),{col_idx},FALSE),0)')
                self._set_cell_formula(comparison_ws, r, i + 3,
                                       f'=IFERROR(VLOOKUP({tx_name_cell},INDIRECT("\'"&$H$2&"\'!A:J"),{col_idx},FALSE),0)')
            # Min, Avg, 90p, Max
            for i, col_idx in enumerate([5, 6, 7, 9], start=8):
                self._set_cell_formula(comparison_ws, r, i,
                                       f'=IFERROR(VLOOKUP({tx_name_cell},INDIRECT("\'"&$H$1&"\'!A:J"),{col_idx},FALSE),0)')
                self._set_cell_formula(comparison_ws, r, i + 1,
                                       f'=IFERROR(VLOOKUP({tx_name_cell},INDIRECT("\'"&$H$2&"\'!A:J"),{col_idx},FALSE),0)')
            # Difference formulas
            self._set_cell_formula(comparison_ws, r, 17, f'=O{r}-N{r}')
            comparison_ws.cell(row=r, column=17).number_format = '0.00'
            self._set_cell_formula(comparison_ws, r, 18, f'=IFERROR((O{r}-N{r})/N{r}, 0)')
            comparison_ws.cell(row=r, column=18).number_format = '0.00%'

        # 6. Apply Conditional Formatting
        last_data_row = comparison_ws.max_row
        if last_data_row > table_start_row:
            self._apply_conditional_formatting(comparison_ws, f"R{table_start_row + 1}:R{last_data_row}", 0.1, 0.2)
            self._apply_conditional_formatting(comparison_ws, f"O{table_start_row + 1}:O{last_data_row}", 1.0, 2.0)

        self.logger.info("Dashboard has been fully rebuilt with headers, data, and formulas.")

    # ... (The other helper methods: _get_or_create_consolidated_workbook, etc., remain the same) ...
    def _get_or_create_consolidated_workbook(self, api: Any, bucket: str, name: str) -> (Workbook, bool):
        try:
            self.logger.info(f"Attempting to download existing consolidated report: {name}")
            file_content = api.download_artifact(bucket, name)
            return load_workbook(filename=io.BytesIO(file_content)), False
        except Exception:
            self.logger.warning(f"Consolidated report not found. Creating a new one.")
            workbook = Workbook();
            workbook.active.title = "Comparison";
            workbook.create_sheet("tests")
            if "Sheet" in workbook.sheetnames: workbook.remove(workbook["Sheet"])
            return workbook, True

    def _add_report_as_hidden_sheet(self, wb: Workbook, report: PerformanceReport) -> str:
        date_obj = self._to_datetime_from_str(report.summary.date_start)
        sheet_name = f"{report.summary.max_user_count}vu_{date_obj.strftime('%Y%m%d_%H%M')}"
        self.logger.info(f"Adding new hidden sheet: {sheet_name}")
        single_report_wb = self.reporter.generate_workbook(report);
        source_ws = single_report_wb.active
        new_ws = wb.create_sheet(title=sheet_name)
        for row in source_ws.iter_rows():
            for cell in row:
                new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font, new_cell.border, new_cell.fill, new_cell.number_format, new_cell.alignment = \
                        copy(cell.font), copy(cell.border), copy(cell.fill), copy(cell.number_format), copy(
                            cell.alignment)
        new_ws.sheet_state = 'hidden'
        return sheet_name

    def _update_tests_index(self, wb: Workbook, sheet_name: str, report: PerformanceReport):
        tests_ws = wb["tests"];
        description = f"{report.test_name} - {report.summary.date_start}"
        tests_ws.append([sheet_name, description]);
        self.logger.info(f"Updated 'tests' index with: {sheet_name}")

    def _prune_old_reports(self, wb: Workbook):
        tests_ws = wb["tests"]
        while tests_ws.max_row > self.history_limit:
            sheet_to_remove = tests_ws.cell(row=1, column=1).value
            self.logger.info(
                f"History limit ({self.history_limit}) exceeded. Removing oldest report sheet: {sheet_to_remove}")
            if sheet_to_remove in wb.sheetnames: wb.remove(wb[sheet_to_remove])
            tests_ws.delete_rows(1)

    def _create_ai_analysis_markdown(self, analysis: Optional[Any]) -> str:
        """Formats the AI analysis object into a readable markdown string."""
        if not analysis or not hasattr(analysis, 'summary'):
            return "# AI Performance Analysis\n\nAI analysis was not available or failed to generate for this comparison."

        md = f"# AI Performance Analysis\n\n"
        md += f"## Overall Summary\n\n{analysis.summary}\n\n"

        if hasattr(analysis, 'key_findings') and analysis.key_findings:
            md += "## Key Findings\n\n"
            for finding in analysis.key_findings:
                md += f"- {finding}\n"

        if hasattr(analysis, 'recommendations') and analysis.recommendations:
            md += "\n## Recommendations\n\n"
            for recommendation in analysis.recommendations:
                md += f"- {recommendation}\n"

        if hasattr(analysis, 'performance_trends') and analysis.performance_trends:
            md += "\n## Performance Trends\n\n"
            for trend_key, trend_value in analysis.performance_trends.items():
                formatted_key = trend_key.replace('_', ' ').title()
                md += f"- **{formatted_key}**: {trend_value}\n"

        if hasattr(analysis, 'risk_assessment') and analysis.risk_assessment:
            md += "\n## Risk Assessment\n\n"
            overall_risk = analysis.risk_assessment.get('overall_risk', 'Unknown')
            md += f"- **Overall Risk Level**: {overall_risk.upper()}\n"

            if 'risk_factors' in analysis.risk_assessment:
                md += "\n### Risk Factors\n"
                for factor in analysis.risk_assessment.get('risk_factors', []):
                    md += f"- {factor}\n"

        if hasattr(analysis, 'confidence_score'):
            md += f"\n---\n*Analysis Confidence Score: {analysis.confidence_score:.0%}*\n"

        return md

    def _create_and_upload_zip_archive(self, api: Any, excel_wb: Workbook, md_content: str, bucket: str,
                                       zip_name: str) -> str:
        self.logger.info(f"Creating in-memory ZIP archive: {zip_name}")
        zip_buffer = io.BytesIO();
        excel_buffer = io.BytesIO();
        excel_wb.save(excel_buffer)
        excel_filename = zip_name.replace(".zip", ".xlsx");
        md_filename = zip_name.replace(".zip", "_analysis.md")
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            zf.writestr(excel_filename, excel_buffer.getvalue());
            zf.writestr(md_filename, md_content.encode('utf-8'))
        zip_bytes = zip_buffer.getvalue();
        self.logger.info(f"In-memory ZIP archive created ({len(zip_bytes)} bytes). Uploading...")
        api.upload_report_from_bytes(zip_bytes, bucket, zip_name)
        return self._generate_download_link(api, bucket, zip_name)

    def _generate_download_link(self, api_wrapper, bucket_name: str, file_name: str) -> str:
        try:
            if hasattr(api_wrapper, 'carrier_client'):
                project_id = api_wrapper.carrier_client.credentials.project_id;
                base_url = api_wrapper.carrier_client.credentials.url.rstrip('/')
            else:
                project_id = api_wrapper._client.credentials.project_id;
                base_url = api_wrapper._client.credentials.url.rstrip('/')
            download_url = f"{base_url}/api/v1/artifacts/artifact/default/{project_id}/{bucket_name}/{file_name}?integration_id=1&is_local=False"
            self.logger.info(f"Successfully generated download link: {download_url}");
            return download_url
        except Exception as e:
            self.logger.warning(f"Could not generate download link: {e}. Providing manual access info.")
            return f"Upload successful. Access file '{file_name}' in bucket '{bucket_name}'."

    def _to_datetime_from_str(self, date_str: str) -> datetime:
        try:
            return datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S')
        except (ValueError, TypeError):
            self.logger.warning(
                f"Could not parse date string '{date_str}'. Falling back to current time.");
            return datetime.now()
