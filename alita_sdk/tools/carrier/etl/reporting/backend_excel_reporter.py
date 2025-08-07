"""
Excel Workbook Generator
Author: Karen Florykian
"""
import logging
import re
from typing import List, Dict, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

from .core.data_models import (
    PerformanceReport,
    TransactionMetrics,
    PerformanceStatus,
    ThresholdConfig,
    PerformanceAnalysisResult
)
from ...utils.utils import ExcelFormattingConfig, DateTimeUtils, ExcelStyleUtils, REPORT_THEME

logger = logging.getLogger(__name__)


class BusinessInsightsGenerator:
    """
    Generates actionable insights from performance data.
    Copilot-friendly: Simple methods for performance analysts to extend.
    """

    def __init__(self, config: ExcelFormattingConfig = None):
        """Initializes the generator with a formatting and thresholds configuration."""
        self.config = config or ExcelFormattingConfig()
        logger.info("BusinessInsightsGenerator initialized")

    def generate_key_insights(self, report: PerformanceReport) -> List[str]:
        """
        Generate key insights. Analysts can extend this by adding new private insight methods.
        """
        insights = [
            self._get_status_insight(report),
            self._get_error_rate_insight(report)
        ]
        insights.extend(self._get_throughput_insights(report))
        insights.extend(self._get_response_time_insights(report))

        # Filter out any None results from insight methods that might not apply
        final_insights = [insight for insight in insights if insight]
        logger.debug(f"Generated {len(final_insights)} insights for report")
        return final_insights

    def _get_status_insight(self, report: PerformanceReport) -> str:
        """Get overall status insight based on build status."""
        if report.build_status == PerformanceStatus.PASSED:
            return "✅ Overall system performance meets defined thresholds"
        return "🔴 System performance degradation detected - immediate attention required"

    def _get_throughput_insights(self, report: PerformanceReport) -> List[str]:
        """Get throughput-related insights."""
        insights = []
        # This threshold could also be moved to config if it varies
        if report.summary.throughput < 10.0:
            slow_transactions = [
                name for name, metrics in report.transactions.items()
                if name != "Total" and metrics.pct95 > self.config.response_time_critical
            ]
            if slow_transactions:
                transaction_list = ", ".join(slow_transactions[:3])
                more_count = len(slow_transactions) - 3
                if more_count > 0:
                    transaction_list += f" and {more_count} more"
                insights.append(f"🔻 Throughput affected by slow transactions: {transaction_list}")
        return insights

    def _get_error_rate_insight(self, report: PerformanceReport) -> str:
        """Get error rate insight."""
        if report.summary.error_rate > self.config.error_rate_warning:
            return f"⚠️ System error rate at {report.summary.error_rate:.2f}% - investigate failed transactions"
        return "✅ Zero error rate - all transactions completing successfully"

    def _get_response_time_insights(self, report: PerformanceReport) -> List[str]:
        """Get response time insights based on critical threshold."""
        insights = []
        critical_transactions = [
            name for name, metrics in report.transactions.items()
            if name != "Total" and metrics.pct95 > self.config.response_time_critical
        ]
        if critical_transactions:
            insights.append(
                f"🔴 Response times > {self.config.response_time_critical / 1000}s by 95th percentile: "
                f"{len(critical_transactions)} transaction(s)"
            )
        return insights

    def filter_insights_by_intent(self, insights: List[str], query: str = "") -> List[str]:
        """
        PLACEHOLDER: Future intent recognition integration.
        For analysts: Add natural language processing here to filter insights.
        Example: "show only critical issues" -> filter insights with 🔴 emoji
        """
        if not query:
            return insights
        # Simple keyword-based filtering as a starting point
        logger.debug(f"Filtering insights with query: '{query}'")
        query_lower = query.lower()
        if "critical" in query_lower or "degradation" in query_lower:
            return [i for i in insights if "🔴" in i]
        if "warning" in query_lower or "investigate" in query_lower:
            return [i for i in insights if "⚠️" in i]
        return insights


# =================================================================================
# 2. EXCEL FORMATTER (Configuration-Driven)
# =================================================================================

class LegacyExcelFormatter:
    """
    Excel formatter using a configuration-driven approach.
    Copilot-friendly: Methods are organized by Excel feature for easy extension.
    """

    def __init__(self, config: ExcelFormattingConfig = None):
        """Initializes the formatter with a configuration object."""
        self.config = config or ExcelFormattingConfig()
        self._setup_fills_and_fonts()
        self.header_mapping = {
            'Transaction': 'name', 'Req, count': 'samples', 'KO, count': 'ko',
            'KO, %': 'error_rate', 'Min, sec': 'min', 'Avg, sec': 'avg',
            '90p, sec': 'pct90', '95p, sec': 'pct95', 'Max, sec': 'max', '📝 Notes': 'notes'
        }
        self.title_mapping = {
            'Users': 'max_user_count', 'Ramp Up, min': 'ramp_up_period', 'Duration, min': 'duration',
            'Think time, sec': 'think_time', 'Start Date, EST': 'date_start', 'End Date, EST': 'date_end',
            'Throughput, req/sec': 'throughput', 'Error rate, %': 'error_rate',
            'Carrier report': 'carrier_report', 'Hypothesis': 'hypothesis', 'Build status': 'build_status',
            'Justification': 'justification', 'Key Insights': 'key_insights',
            'Overall system performance': 'overall_performance'
        }
        logger.info("LegacyExcelFormatter initialized with configuration")

    def _setup_fills_and_fonts(self):
        """Creates openpyxl style objects from the configuration hex colors."""
        self.red_fill = PatternFill(
            start_color=self.config.colors['critical'].replace('#', ''),
            end_color=self.config.colors['critical'].replace('#', ''), fill_type='solid'
        )
        self.green_fill = PatternFill(
            start_color=self.config.colors['good'].replace('#', ''),
            end_color=self.config.colors['good'].replace('#', ''), fill_type='solid'
        )
        self.yellow_fill = PatternFill(
            start_color=self.config.colors['warning'].replace('#', ''),
            end_color=self.config.colors['warning'].replace('#', ''), fill_type='solid'
        )
        # Font colors for status indicators
        self.green_font_color = self.config.colors['excellent'].replace('#', '')
        self.red_font_color = self.config.colors['critical_font'].replace('#', '')

    def apply_title_section_formatting(self, ws: Worksheet, report: PerformanceReport, insights: List[str]):
        """Apply title section formatting using data from the report."""
        report_data = {
            'max_user_count': getattr(report.summary, 'max_user_count', 'N/A'),
            'ramp_up_period': getattr(report.summary, 'ramp_up_period', 'N/A'),
            'duration': getattr(report.summary, 'duration', 'N/A'),
            'think_time': getattr(report.summary, 'think_time', 'N/A'),
            'date_start': report.summary.date_start.strftime('%Y-%m-%d %H:%M:%S')
            if hasattr(report.summary.date_start, 'strftime') else report.summary.date_start or 'N/A',
            'date_end': report.summary.date_end.strftime('%Y-%m-%d %H:%M:%S')
            if hasattr(report.summary.date_end, 'strftime') else report.summary.date_end or 'N/A',
            'throughput': report.summary.throughput,
            'error_rate': report.summary.error_rate / 100.0,
            'carrier_report': getattr(report, 'carrier_report_url', 'N/A'),
            'hypothesis': getattr(report.summary, 'hypothesis', 'Validate against SLAs'),
            'build_status': report.build_status.value,
            'justification': report.analysis_summary or "Analysis not provided.",
            'key_insights': '\n'.join(insights),
            'overall_performance': 'HEALTHY' if report.build_status == PerformanceStatus.PASSED else 'DEGRADED'
        }

        border_style = Side(border_style="thin", color="040404")
        border = Border(top=border_style, left=border_style, right=border_style, bottom=border_style)
        title_fill = PatternFill("solid", fgColor='00CDEBEA')

        for i, (title_name, data_key) in enumerate(self.title_mapping.items(), 1):
            title_cell = ws.cell(row=i, column=1, value=title_name)
            title_cell.font = Font(bold=True, color='00291A75')
            title_cell.fill = title_fill
            title_cell.alignment = Alignment(horizontal="left", vertical="center")
            title_cell.border = border

            value_cell = ws.cell(row=i, column=2, value=report_data[data_key])
            value_cell.fill = title_fill
            value_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            value_cell.border = border

            if data_key == 'error_rate':
                value_cell.number_format = '0.00%'
            elif data_key == 'build_status':
                color = self.green_font_color if report_data[data_key] == 'PASSED' else self.red_font_color
                value_cell.font = Font(bold=True, color=color)
            elif data_key in ['justification', 'key_insights']:
                value_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                ws.row_dimensions[i].height = max(50, 15 * (str(report_data[data_key]).count('\n') + 1))

            ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=len(self.header_mapping))

        return len(self.title_mapping) + 1

    def apply_transaction_table_formatting(self, ws: Worksheet, report: PerformanceReport, start_row: int,
                                           thresholds: Dict[str, float]):
        """Apply transaction table formatting with config-driven conditional formatting."""
        header_row = start_row
        for col_idx, header_text in enumerate(self.header_mapping.keys(), 1):
            cell = ws.cell(row=header_row, column=col_idx, value=header_text)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill("solid", fgColor='007FD5D8')
            border_style = Side(border_style="thin", color="040404")
            cell.border = Border(top=border_style, left=border_style, right=border_style, bottom=border_style)

        current_row = header_row + 1
        sorted_transactions = sorted([name for name in report.transactions.keys() if name != "Total"])
        if "Total" in report.transactions:
            sorted_transactions.append("Total")

        for tx_name in sorted_transactions:
            metrics = report.transactions[tx_name]
            for col_idx, (header_name, data_key) in enumerate(self.header_mapping.items(), 1):
                cell_value = self._generate_validation_notes(metrics, thresholds) if data_key == 'notes' else getattr(
                    metrics, data_key, 'N/A')
                cell = ws.cell(row=current_row, column=col_idx, value=cell_value)
                border_style = Side(border_style="thin", color="040404")
                cell.border = Border(top=border_style, left=border_style, right=border_style, bottom=border_style)

                if data_key in ['samples', 'ko', 'min', 'avg', 'pct90', 'pct95', 'max'] and isinstance(cell_value,
                                                                                                       (int, float)):
                    cell.value = round(cell_value / 1000, 3)
                elif data_key == 'error_rate' and isinstance(cell_value, (int, float)):
                    cell.number_format = '0.00%'
                    cell.value = cell_value / 100.0

                if tx_name == "Total":
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor='007FD5D8')

            current_row += 1

        content_end_col = get_column_letter(len(self.header_mapping))
        self._apply_conditional_formatting(ws, header_row + 1, current_row - 1, thresholds)
        self._apply_legacy_table_features(ws, f"A{header_row}", f"{content_end_col}{current_row - 1}", header_row)
        return current_row

    def _generate_validation_notes(self, metrics: TransactionMetrics, thresholds: Dict[str, float]) -> str:
        """Generate validation notes based on response time thresholds from config."""
        logger.debug(f"Generating validation notes for {metrics.name}")
        rt_threshold = thresholds.get('response_time', self.config.response_time_warning)
        threshold_seconds = rt_threshold / 1000
        p95_seconds = metrics.pct95 / 1000 if hasattr(metrics, 'pct95') else 0

        if p95_seconds <= threshold_seconds:
            return "✅ Within SLA"
        else:
            pct_over = ((p95_seconds - threshold_seconds) / threshold_seconds) * 100
            return f"⚠️ {pct_over:.1f}% over SLA"

    def _apply_conditional_formatting(self, ws: Worksheet, start_row: int, end_row: int, thresholds: Dict[str, float]):
        """Apply conditional formatting using configuration-driven rules."""
        rt_threshold = thresholds.get('response_time', self.config.response_time_warning) / 1000  # Convert to seconds
        logger.debug(f"Applying conditional formatting with RT threshold: {rt_threshold}s")

        time_columns = [5, 6, 7, 8, 9]  # Columns E through I
        for col_idx in time_columns:
            col_letter = get_column_letter(col_idx)
            range_coords = f"{col_letter}{start_row}:{col_letter}{end_row}"

            ws.conditional_formatting.add(range_coords,
                                          CellIsRule(operator='lessThan', formula=[rt_threshold], fill=self.green_fill))
            ws.conditional_formatting.add(range_coords, CellIsRule(operator='greaterThan', formula=[rt_threshold * 1.5],
                                                                   fill=self.red_fill))
            ws.conditional_formatting.add(range_coords,
                                          CellIsRule(operator='between', formula=[rt_threshold, rt_threshold * 1.5],
                                                     fill=self.yellow_fill))

    def _apply_legacy_table_features(self, ws: Worksheet, content_start_ref: str, content_end_ref: str,
                                     header_row: int):
        """Apply legacy table features: auto-filter, freeze panes, and column widths."""
        logger.debug(f"Applying table features to range {content_start_ref}:{content_end_ref}")
        ws.auto_filter.ref = f"{content_start_ref}:{content_end_ref}"
        ws.freeze_panes = ws[f'B{header_row + 1}']
        for i, header in enumerate(self.header_mapping.keys(), 1):
            col_letter = get_column_letter(i)
            if header == "Transaction":
                max_length = max(len(str(cell.value)) for cell in ws[col_letter])
                ws.column_dimensions[col_letter].width = max(20, max_length + 5)
            else:
                ws.column_dimensions[col_letter].width = len(header) + 5


class ExcelReporter:
    """
    Production Excel reporter with configuration-driven formatting.
    Copilot-friendly: Clear separation of concerns for easy extension by analysts.
    """

    def __init__(self, config: ExcelFormattingConfig = None):
        """Initializes the reporter and its components with a shared configuration."""
        self.config = config or ExcelFormattingConfig()
        self.formatter = LegacyExcelFormatter(self.config)
        self.insights_generator = BusinessInsightsGenerator(self.config)
        logger.info("ExcelReporter initialized with configuration")

    def generate_workbook(self, report: PerformanceReport, analyst_query: str = "") -> Workbook:
        """
        Generate a production-ready Excel workbook.
        Optionally filters insights based on an analyst's natural language query.
        """
        logger.info(f"Generating Excel report for build status: {report.build_status.value}")
        wb = Workbook()
        ws = wb.active
        ws.title = "Test results"

        insights = self.insights_generator.generate_key_insights(report)
        if analyst_query:
            insights = self.insights_generator.filter_insights_by_intent(insights, analyst_query)

        thresholds = self._extract_thresholds(report.thresholds)
        next_row = self.formatter.apply_title_section_formatting(ws, report, insights)
        self.formatter.apply_transaction_table_formatting(ws, report, next_row, thresholds)

        logger.info("Excel report generated successfully.")
        return wb

    def generate_workbook_with_intent(self, report: PerformanceReport, query: str = "") -> Workbook:
        """
        PLACEHOLDER: Generate workbook filtered by analyst intent/query.
        This provides a clear entry point for future NLP-driven report customization.
        """
        logger.info(f"Generating intent-driven workbook with query: '{query}'")
        return self.generate_workbook(report, analyst_query=query)

    def _extract_thresholds(self, threshold_configs: List[ThresholdConfig]) -> Dict[str, float]:
        """Extract threshold values, using config defaults if not provided."""
        thresholds = {}
        if threshold_configs:
            for config in threshold_configs:
                thresholds[config.target] = config.threshold_value

        # Apply defaults from ExcelFormattingConfig if not explicitly set in the report
        thresholds.setdefault('response_time', self.config.response_time_warning)
        thresholds.setdefault('error_rate', self.config.error_rate_warning)

        logger.debug(f"Using thresholds for formatting: {thresholds}")
        return thresholds

    def _update_tests_overview_sheet(self, wb: Workbook, new_sheet_name: str, report: PerformanceReport):
        """Update or create a 'Tests Overview' sheet with a hyperlinked summary of the new report."""
        sheet_title = "Tests Overview"
        if sheet_title not in wb.sheetnames:
            ws = wb.create_sheet(sheet_title, 0)
            ws.append(["Test Date", "Report Sheet", "Build Status", "Throughput (req/s)", "P95 Response Time (ms)",
                       "Error Rate", "Key Issues"])
        else:
            ws = wb[sheet_title]

        key_issues = "None"
        if report.build_status != PerformanceStatus.PASSED:
            failed_insights = [i for i in self.insights_generator.generate_key_insights(report) if
                               "🔴" in i or "⚠️" in i]
            key_issues = ", ".join(i.split(' - ')[0] for i in failed_insights) or "Degradation detected"

        new_row = [
            report.summary.date_start.strftime('%Y-%m-%d %H:%M'),
            f'=HYPERLINK("#{new_sheet_name}!A1", "{new_sheet_name}")',
            report.build_status.value,
            report.summary.throughput,
            report.transactions.get("Total",
                                    TransactionMetrics(name="Total", samples=0, ko=0, avg=0, min=0, max=0, pct90=0,
                                                       pct95=0, pct99=0, throughput=0.0, received_kb_per_sec=0.0,
                                                       sent_kb_per_sec=0.0, total=0)).pct95,
            report.summary.error_rate / 100.0,
            key_issues
        ]
        ws.append(new_row)

        row_num = ws.max_row
        ws.cell(row=row_num, column=2).font = Font(bold=True, underline="single", color='00291A75')
        ws.cell(row=row_num, column=6).number_format = '0.00%'
        status_cell = ws.cell(row=row_num, column=3)
        if report.build_status == PerformanceStatus.PASSED:
            status_cell.font = Font(bold=True, color=self.formatter.green_font_color)
        else:
            status_cell.font = Font(bold=True, color=self.formatter.red_font_color)

    def update_report_with_new_sheet(self, workbook_path: str, new_report: PerformanceReport) -> bool:
        """Update an existing workbook with a new sheet containing the new report data."""
        try:
            wb = load_workbook(workbook_path)
            logger.info(f"Loaded existing workbook from '{workbook_path}'.")
        except FileNotFoundError:
            logger.warning(f"Workbook not found at '{workbook_path}'. Creating new workbook.")
            wb = Workbook()
            if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])

        sheet_name = new_report.summary.date_start.strftime('%Y-%m-%d_%H%M')
        if sheet_name in wb.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' already exists. Replacing it.")
            wb.remove(wb[sheet_name])

        ws_new = wb.create_sheet(sheet_name)

        insights = self.insights_generator.generate_key_insights(new_report)
        thresholds = self._extract_thresholds(new_report.thresholds)

        next_row = self.formatter.apply_title_section_formatting(ws_new, new_report, insights)
        self.formatter.apply_transaction_table_formatting(ws_new, new_report, next_row, thresholds)

        self._update_tests_overview_sheet(wb, sheet_name, new_report)

        wb.save(workbook_path)
        logger.info(f"Successfully updated workbook saved to '{workbook_path}'.")
        return True


class ComparisonReporter:
    """
    Builds a sophisticated, multi-sheet comparison workbook that integrates the
    best features of the legacy system with a modern, maintainable, and DRY architecture.
    """

    def __init__(self, config: ExcelFormattingConfig = None):
        self.config = config or ExcelFormattingConfig()
        self.formatter = LegacyExcelFormatter(self.config)
        self.logger = logging.getLogger(__name__)

    def generate_comparison_workbook(self, reports: List[PerformanceReport],
                                     ai_analysis: Optional[PerformanceAnalysisResult] = None) -> Workbook:
        """
        Main method to generate the complete, multi-sheet workbook object.
        """
        if not reports or len(reports) < 2:
            raise ValueError(f"Need at least 2 reports for comparison, got {len(reports) if reports else 0}")

        self.logger.info(f"Generating interactive comparison workbook with {len(reports)} reports...")
        wb = self._create_base_workbook()
        tests_ws = wb["Tests"]

        for report in reports:
            sheet_name = self._generate_sheet_name(report)
            if sheet_name not in wb.sheetnames:
                self._add_signle_report_sheet(wb, report, sheet_name)
                self._update_tests_registry(tests_ws, sheet_name, report)

        self._build_comparison_dashboard(wb, ai_analysis)

        self.logger.info("Comparison workbook object created successfully.")
        return wb

    # --- Core Workbook and Sheet Management ---

    def _create_base_workbook(self) -> Workbook:
        """Creates a new workbook with the required base sheets."""
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        wb.create_sheet("Comparison", 0)
        wb.create_sheet("Tests", 1)
        # wb["Tests"].sheet_state = 'hidden'
        return wb

    def _generate_sheet_name(self, report: PerformanceReport) -> str:
        """Generates a unique, human-readable sheet name for a report."""
        summary = report.summary
        date_obj = DateTimeUtils.parse_datetime(summary.date_start)
        user_count = summary.max_user_count
        test_name = re.sub(r'[\\/*?:\[\]]', '', getattr(report, 'test_name', 'test'))[:15]
        return f"{user_count}vu_{test_name}_{date_obj.strftime('%m%d_%H%M')}"

    def _add_signle_report_sheet(self, wb: Workbook, report: PerformanceReport, sheet_name: str):
        """Uses the existing LegacyExcelFormatter to create a data sheet."""
        ws = wb.create_sheet(sheet_name)
        insights = BusinessInsightsGenerator(self.config).generate_key_insights(report)
        thresholds = {
            'response_time': self.config.response_time_warning,
            'error_rate': self.config.error_rate_warning
        }

        next_row = self.formatter.apply_title_section_formatting(ws, report, insights)
        self.formatter.apply_transaction_table_formatting(ws, report, next_row, thresholds)

        # ws.sheet_state = 'hidden'
        # self.logger.debug(f"Added hidden data sheet: {sheet_name}")

    def _update_tests_registry(self, tests_ws: Worksheet, sheet_name: str, report: PerformanceReport):
        """Adds an entry to the 'Tests' registry sheet."""
        description = f"{report.summary.max_user_count} Users - {getattr(report, 'test_name', 'N/A')}"
        tests_ws.append([sheet_name, description])

    # --- Dashboard Orchestration ---

    def _build_comparison_dashboard(self, wb: Workbook, ai_analysis: Optional[PerformanceAnalysisResult]):
        """Orchestrates the creation of the main 'Comparison' dashboard."""
        ws = wb["Comparison"]

        self._create_dropdowns(ws, wb["Tests"])
        self._create_summary_section(ws)
        last_row = self._create_transaction_table(ws, wb)

        if ai_analysis:
            self._add_ai_analysis_sheet(wb, ai_analysis)

        self._apply_conditional_formatting(ws, last_row)
        ExcelStyleUtils.auto_adjust_column_width(ws, max_width=40)

    # --- Interactive Features ---

    def _create_dropdowns(self, ws: Worksheet, tests_ws: Worksheet):
        """Creates interactive dropdowns and sets smart defaults."""
        # Dropdown labels
        label1_cell = ws.cell(row=1, column=1, value="Select Test 1:")
        label2_cell = ws.cell(row=2, column=1, value="Select Test 2:")
        label1_cell.font = REPORT_THEME.FONT_UI_HEADER
        label2_cell.font = REPORT_THEME.FONT_UI_HEADER

        # Create data validation for dropdowns
        if tests_ws.max_row > 0:
            data_val = DataValidation(
                type="list",
                formula1=f'Tests!$A$1:$A${tests_ws.max_row}',
                allow_blank=False
            )
            ws.add_data_validation(data_val)
            data_val.add(ws["B1"])
            data_val.add(ws["B2"])

            # Apply dropdown styling
            dropdown_cells = [ws["B1"], ws["B2"]]
            for cell in dropdown_cells:
                cell.fill = REPORT_THEME.FILL_DROPDOWN
                cell.font = REPORT_THEME.FONT_UI_HEADER

        # Set smart defaults
        self._set_smart_dropdown_defaults(ws, tests_ws)

    def _set_smart_dropdown_defaults(self, ws: Worksheet, tests_ws: Worksheet):
        """Pre-selects the latest test and finds the best matching baseline."""
        if tests_ws.max_row < 1:
            return

        # Default comparison is always the latest test
        latest_test_name = tests_ws.cell(row=tests_ws.max_row, column=1).value
        latest_test_desc = tests_ws.cell(row=tests_ws.max_row, column=2).value
        ws["B2"].value = latest_test_name

        # Find the best baseline
        baseline_name = None
        if latest_test_desc:
            match = re.search(r"(\d+)\s*Users", str(latest_test_desc))
            if match:
                latest_vu = match.group(1)
                # Search backwards for a test with the same user count
                for row in range(tests_ws.max_row - 1, 0, -1):
                    desc = tests_ws.cell(row=row, column=2).value
                    if desc and f"{latest_vu} Users" in desc:
                        baseline_name = tests_ws.cell(row=row, column=1).value
                        break

        # If no baseline found, use the previous test
        if not baseline_name and tests_ws.max_row >= 2:
            baseline_name = tests_ws.cell(row=tests_ws.max_row - 1, column=1).value

        ws["B1"].value = baseline_name or latest_test_name

    # --- Dynamic Content Generation ---

    def _create_summary_section(self, ws: Worksheet):
        """Builds the summary table with INDIRECT formulas."""
        headers = ["Users", "Ramp up, sec", "Duration, min", "Think time, sec",
                   "Start Date, BST", "End Date, BST", "Throughput, req/sec", "Error rate, %"]

        for i, header in enumerate(headers, 4):
            # Header
            header_cell = ws.cell(row=i, column=1, value=header)
            header_cell.font = REPORT_THEME.FONT_SUMMARY_LABEL

            # Test 1 formula (baseline)
            test1_cell = ws.cell(row=i, column=2,
                                 value=f"=IF($B$1=\"\",\"\",INDIRECT(\"'\"&$B$1&\"'!B{i - 3}\"))")
            ExcelStyleUtils.apply_data_style(test1_cell)

            # Test 2 formula (comparison)
            test2_cell = ws.cell(row=i, column=3,
                                 value=f"=IF($B$2=\"\",\"\",INDIRECT(\"'\"&$B$2&\"'!B{i - 3}\"))")
            ExcelStyleUtils.apply_data_style(test2_cell)

    def _create_transaction_table(self, ws: Worksheet, wb: Workbook) -> int:
        """Builds the detailed, multi-column transaction comparison table."""
        all_transactions = self._get_all_transactions(wb)
        if not all_transactions:
            return ws.max_row

        start_row = ws.max_row + 3

        # Build the exact table structure from your screenshot
        self._create_transaction_headers(ws, start_row)
        self._populate_transaction_data(ws, all_transactions, start_row + 2)

        return ws.max_row

    def _create_transaction_headers(self, ws: Worksheet, start_row: int):
        """Create the exact header structure matching the screenshot."""
        main_headers = [
            ('Transaction', 1, 1),
            ('Req. count', 2, 2),
            ('KO, count', 4, 2),
            ('Min, sec', 6, 2),
            ('Avg, sec', 8, 2),
            ('90p, sec', 10, 2),
            ('95p, sec', 12, 2),
            ('Difference, 90p', 14, 4),  # This spans 4 columns
            ('Max, sec', 18, 2)
        ]

        # Create main headers with merging
        for header_text, start_col, span in main_headers:
            cell = ws.cell(row=start_row, column=start_col, value=header_text)
            REPORT_THEME.apply_header_style(cell)

            if span > 1:
                ws.merge_cells(start_row=start_row, start_column=start_col,
                               end_row=start_row, end_column=start_col + span - 1)

        # Sub headers (row 2)
        sub_headers = [
            (2, 'test1'), (3, 'test2'),  # Req. count
            (4, 'test1'), (5, 'test2'),  # KO, count
            (6, 'test1'), (7, 'test2'),  # Min, sec
            (8, 'test1'), (9, 'test2'),  # Avg, sec
            (10, 'test1'), (11, 'test2'),  # 90p, sec
            (12, 'test1'), (13, 'test2'),  # 95p, sec
            (14, 'Diff, sec'), (15, 'Diff, %'), (16, 'Diff, sec'), (17, 'Diff, %'),  # Difference columns
            (18, 'test1'), (19, 'test2')  # Max, sec
        ]

        for col, text in sub_headers:
            cell = ws.cell(row=start_row + 1, column=col, value=text)
            REPORT_THEME.apply_header_style(cell)

    def _populate_transaction_data(self, ws: Worksheet, transactions: List[str], start_row: int):
        """Populate transaction data with proper formulas matching the data structure."""
        for row_idx, tx_name in enumerate(transactions, start_row):
            # Transaction name
            ws.cell(row=row_idx, column=1, value=tx_name)

            metric_columns = [
                (2, 3, 2),  # Req. count - column B
                (4, 5, 3),  # KO, count - column C
                (6, 7, 5),  # Min, sec - column E (converted from ms to sec)
                (8, 9, 6),  # Avg, sec - column F
                (10, 11, 7),  # 90p, sec - column G
                (12, 13, 8),  # 95p, sec - column H
                (18, 19, 9)  # Max, sec - column I
            ]

            for test1_col, test2_col, source_col in metric_columns:
                # Test1 formula (baseline)
                test1_formula = f"=IFERROR(IF($B$1=\"\",\"\",VLOOKUP($A{row_idx},INDIRECT(\"'\"&$B$1&\"'!$A:$J\"),{source_col},FALSE)),\"\")"
                ws.cell(row=row_idx, column=test1_col, value=test1_formula)

                # Test2 formula (comparison)
                test2_formula = f"=IFERROR(IF($B$2=\"\",\"\",VLOOKUP($A{row_idx},INDIRECT(\"'\"&$B$2&\"'!$A:$J\"),{source_col},FALSE)),\"\")"
                ws.cell(row=row_idx, column=test2_col, value=test2_formula)

            # Special handling for 90p difference columns (columns 14-17)
            test1_90p_col = get_column_letter(10)  # 90p test1
            test2_90p_col = get_column_letter(11)  # 90p test2

            # Difference in seconds (absolute)
            diff_sec_formula = f"=IF(AND({test1_90p_col}{row_idx}<>\"\",{test2_90p_col}{row_idx}<>\"\"),{test2_90p_col}{row_idx}-{test1_90p_col}{row_idx},\"\")"
            ws.cell(row=row_idx, column=14, value=diff_sec_formula)

            # Difference in percentage
            diff_pct_formula = f"=IF(AND({test1_90p_col}{row_idx}>0,{test2_90p_col}{row_idx}<>\"\"),({test2_90p_col}{row_idx}-{test1_90p_col}{row_idx})/{test1_90p_col}{row_idx},\"\")"
            pct_cell = ws.cell(row=row_idx, column=15, value=diff_pct_formula)
            pct_cell.number_format = '0.00%'

            # Additional difference columns (95p) - columns 16-17
            test1_95p_col = get_column_letter(12)  # 95p test1
            test2_95p_col = get_column_letter(13)  # 95p test2

            diff_95p_sec_formula = f"=IF(AND({test1_95p_col}{row_idx}<>\"\",{test2_95p_col}{row_idx}<>\"\"),{test2_95p_col}{row_idx}-{test1_95p_col}{row_idx},\"\")"
            ws.cell(row=row_idx, column=16, value=diff_95p_sec_formula)

            diff_95p_pct_formula = f"=IF(AND({test1_95p_col}{row_idx}>0,{test2_95p_col}{row_idx}<>\"\"),({test2_95p_col}{row_idx}-{test1_95p_col}{row_idx})/{test1_95p_col}{row_idx},\"\")"
            pct_95_cell = ws.cell(row=row_idx, column=17, value=diff_95p_pct_formula)
            pct_95_cell.number_format = '0.00%'

    def _get_all_transactions(self, wb: Workbook) -> List[str]:
        """Gets a sorted, unique list of all transaction names from data sheets."""
        transactions = set()
        for sheet in wb.worksheets:
            if sheet.title not in ["Comparison", "Tests", "AI Analysis"] and sheet.sheet_state != 'hidden':
                # Look for transaction data starting from a reasonable row (after headers)
                for row in range(12, min(sheet.max_row + 1, 100)):  # Limit search
                    cell_val = sheet.cell(row=row, column=1).value
                    if cell_val and isinstance(cell_val,
                                               str) and cell_val != "Total" and cell_val != "Transaction" and cell_val != "Justification" and cell_val != "Key Insights" and cell_val != "Overall system performance":
                        transactions.add(cell_val)
        return sorted(list(transactions))

    def _add_ai_analysis_sheet(self, wb: Workbook, analysis: PerformanceAnalysisResult):
        """Adds the AI analysis to its own dedicated sheet."""
        ws = wb.create_sheet("AI Analysis")
        title_cell = ws.cell(row=1, column=1, value="AI Performance Analysis")
        title_cell.font = REPORT_THEME.FONT_UI_BOLD

        current_row = 3

        # Executive Summary
        summary_header = ws.cell(row=current_row, column=1, value="Executive Summary")
        summary_header.font = REPORT_THEME.FONT_UI_HEADER

        summary_content = ws.cell(row=current_row + 1, column=1, value=analysis.summary)
        summary_content.alignment = REPORT_THEME.ALIGN_JUSTIFY_TOP

        current_row += 3

        # Key Findings
        findings_header = ws.cell(row=current_row, column=1, value="Key Findings")
        findings_header.font = REPORT_THEME.FONT_UI_HEADER

        for finding in analysis.key_findings:
            current_row += 1
            finding_cell = ws.cell(row=current_row, column=1, value=f"• {finding}")
            finding_cell.alignment = REPORT_THEME.ALIGN_LEFT_CENTER

        current_row += 2

        # Recommendations
        rec_header = ws.cell(row=current_row, column=1, value="Recommendations")
        rec_header.font = REPORT_THEME.FONT_UI_HEADER

        for rec in analysis.recommendations:
            current_row += 1
            rec_cell = ws.cell(row=current_row, column=1, value=f"• {rec}")
            rec_cell.alignment = REPORT_THEME.ALIGN_LEFT_CENTER

        ExcelStyleUtils.auto_adjust_column_width(ws, max_width=100)

    def _apply_conditional_formatting(self, ws: Worksheet, last_data_row: int):
        """Applies conditional formatting to difference columns for visual comparison."""
        # Apply conditional formatting to percentage difference columns (15 and 17)
        diff_columns = [15, 17]  # Diff % columns for 90p and 95p

        for col_num in diff_columns:
            col_letter = get_column_letter(col_num)
            cell_range = f'{col_letter}3:{col_letter}{last_data_row}'

            # Improvement (green for negative values - better performance)
            improvement_rule = CellIsRule(
                operator='lessThan',
                formula=[-0.05],  # 5% improvement threshold
                fill=PatternFill(start_color=self.config.colors['good'], fill_type='solid')
            )

            # Degradation (red for positive values - worse performance)
            degradation_rule = CellIsRule(
                operator='greaterThan',
                formula=[0.05],  # 5% degradation threshold
                fill=PatternFill(start_color=self.config.colors['critical'], fill_type='solid')
            )

            ws.conditional_formatting.add(cell_range, improvement_rule)
            ws.conditional_formatting.add(cell_range, degradation_rule)

        # Apply conditional formatting to absolute response time columns
        # Highlight high response times in test1 and test2 columns
        response_time_columns = [10, 11, 12, 13]  # 90p and 95p columns

        for col_num in response_time_columns:
            col_letter = get_column_letter(col_num)
            cell_range = f'{col_letter}3:{col_letter}{last_data_row}'

            # Warning for response times > 1 second
            warning_rule = CellIsRule(
                operator='greaterThan',
                formula=[1.0],
                fill=PatternFill(start_color=self.config.colors['warning'], fill_type='solid')
            )

            # Critical for response times > 2 seconds
            critical_rule = CellIsRule(
                operator='greaterThan',
                formula=[2.0],
                fill=PatternFill(start_color=self.config.colors['critical'], fill_type='solid')
            )

            ws.conditional_formatting.add(cell_range, warning_rule)
            ws.conditional_formatting.add(cell_range, critical_rule)
