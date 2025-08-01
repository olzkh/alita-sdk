"""
Consolidated UI Excel Reporter

Author: Karen Florykian
"""
import logging
from typing import List, Union
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

from ..reporting.core.data_models import UIPerformanceReport, UIMetrics
from ..transformers.ui_transformers import UITransformResult

logger = logging.getLogger(__name__)


class UIExcelReporter:
    """
    UI Excel reporter that generates workbooks from UIPerformanceReport objects.
    Handles both single UIPerformanceReport and UITransformResult with multiple reports.
    """

    def __init__(self):
        logger.info("UIExcelReporter initialized")

    def generate_workbook(self, transform_result: Union[UIPerformanceReport, 'UITransformResult']) -> Workbook:
        """
        Generate Excel workbook from transformation result.
        Handles both UIPerformanceReport and UITransformResult objects.
        """
        if not transform_result:
            raise ValueError("Transform result is required")

        # Case 1: Direct UIPerformanceReport object
        if isinstance(transform_result, UIPerformanceReport):
            logger.info("Processing single UIPerformanceReport")
            return self._generate_from_single_report(transform_result)

        # Case 2: UITransformResult with worksheets_data dictionary
        elif hasattr(transform_result, 'worksheets_data'):
            logger.info("Processing UITransformResult with worksheets_data")
            return self._generate_from_transform_result(transform_result)

        else:
            raise ValueError(f"Unexpected transform_result type: {type(transform_result)}")

    def _generate_from_single_report(self, ui_report: UIPerformanceReport) -> Workbook:
        """Generate workbook from a single UIPerformanceReport."""
        wb = Workbook()
        ws = wb.active
        ws.title = self._sanitize_sheet_name(ui_report.report_name)

        # Create pivot table from UIMetrics
        df_pivot = self._create_pivot_from_ui_metrics(ui_report.worksheets_data)

        # Write to worksheet
        self._write_pivot_to_worksheet(ws, df_pivot)
        self._apply_formatting(ws, df_pivot)

        logger.info(f"Generated workbook for single report: {ui_report.report_name}")
        return wb

    def _generate_from_transform_result(self, transform_result: 'UITransformResult') -> Workbook:
        """Generate workbook from UITransformResult with multiple reports."""
        wb = Workbook()

        # Remove default sheet
        if wb.worksheets:
            wb.remove(wb.active)

        worksheets_data = transform_result.worksheets_data

        if not worksheets_data:
            raise ValueError("No worksheets data provided")

        # Process each worksheet
        for worksheet_name, ui_report in worksheets_data.items():
            if isinstance(ui_report, UIPerformanceReport):
                ws = wb.create_sheet(title=self._sanitize_sheet_name(worksheet_name))

                # Create pivot table from UIMetrics
                df_pivot = self._create_pivot_from_ui_metrics(ui_report.worksheets_data)

                # Write to worksheet
                self._write_pivot_to_worksheet(ws, df_pivot)
                self._apply_formatting(ws, df_pivot)
            else:
                logger.warning(f"Skipping non-UIPerformanceReport object for worksheet: {worksheet_name}")

        if not wb.worksheets:
            raise ValueError("No worksheets were created")

        logger.info(f"Generated workbook with {len(wb.worksheets)} worksheets")
        return wb

    def _create_pivot_from_ui_metrics(self, ui_metrics_list: List[UIMetrics]) -> pd.DataFrame:
        """Create pivot DataFrame from list of UIMetrics."""
        if not ui_metrics_list:
            raise ValueError("No UI metrics provided")

        # Convert UIMetrics to DataFrame rows
        data_rows = []
        for metric in ui_metrics_list:
            if hasattr(metric, 'step_name'):
                data_rows.append({
                    "Step name": metric.step_name,
                    "Audit": metric.audit,
                    "Numeric Value": metric.numeric_value
                })
            else:
                logger.warning(f"Skipping non-UIMetrics object: {type(metric)}")

        if not data_rows:
            raise ValueError("No valid metrics data found")

        # Create pivot table
        df = pd.DataFrame(data_rows)
        df_pivot = df.pivot_table(
            index="Step name",
            columns="Audit",
            values="Numeric Value",
            aggfunc='mean'
        )
        df_pivot = df_pivot.fillna('')

        return df_pivot

    def _sanitize_sheet_name(self, name: str) -> str:
        """Sanitize sheet name to meet Excel requirements."""
        # Excel sheet name must be <= 31 chars and not contain certain characters
        invalid_chars = [':', '\\', '/', '?', '*', '[', ']']
        sanitized = name
        for char in invalid_chars:
            sanitized = sanitized.replace(char, '_')
        return sanitized[:31]

    def _write_pivot_to_worksheet(self, ws, df_pivot):
        """Write pivot DataFrame to worksheet."""
        # Write column headers (audit names)
        ws.cell(row=1, column=1, value="Step name")
        for col_idx, col_name in enumerate(df_pivot.columns, 2):
            ws.cell(row=1, column=col_idx, value=col_name)

        # Write data rows
        for row_idx, (step_name, row_data) in enumerate(df_pivot.iterrows(), 2):
            ws.cell(row=row_idx, column=1, value=step_name)
            for col_idx, value in enumerate(row_data, 2):
                if value != '':  # Only write non-empty values
                    ws.cell(row=row_idx, column=col_idx, value=float(value))

    def _apply_formatting(self, ws, df_pivot):
        """Apply Excel formatting matching legacy style."""
        # Header formatting
        header_fill = PatternFill(start_color="7FD5D8", end_color="7FD5D8", fill_type="solid")
        header_font = Font(bold=True)

        for cell in ws[1]:  # First row
            cell.fill = header_fill
            cell.font = header_font

        # Step name column alignment
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
            for cell in row:
                cell.alignment = Alignment(horizontal='left')

        # Apply conditional formatting based on audit type
        for col_idx, col_name in enumerate(df_pivot.columns, 2):
            column_letter = get_column_letter(col_idx)

            # CLS thresholds
            if "Cumulative Layout Shift" in col_name or "CLS" in col_name:
                self._apply_conditional_formatting(ws, column_letter, [0.1, 0.25])
            # FCP thresholds (in milliseconds)
            elif "First Contentful Paint" in col_name or "FCP" in col_name:
                self._apply_conditional_formatting(ws, column_letter, [1800, 3000])
            # LCP thresholds (in milliseconds)
            elif "Largest Contentful Paint" in col_name or "LCP" in col_name:
                self._apply_conditional_formatting(ws, column_letter, [2500, 4000])
            # TBT thresholds (in milliseconds)
            elif "Total Blocking Time" in col_name or "TBT" in col_name:
                self._apply_conditional_formatting(ws, column_letter, [200, 600])
            # Speed Index thresholds (in milliseconds)
            elif "Speed Index" in col_name:
                self._apply_conditional_formatting(ws, column_letter, [3400, 5800])
            # TTI thresholds (in milliseconds)
            elif "Time to Interactive" in col_name or "TTI" in col_name:
                self._apply_conditional_formatting(ws, column_letter, [3800, 7300])
        self._apply_borders(ws)
        self._auto_adjust_columns(ws)

    def _apply_conditional_formatting(self, ws, column_letter: str, thresholds: List[float]):
        """Apply color-based conditional formatting to a column."""
        # Green for good performance
        ws.conditional_formatting.add(
            f'{column_letter}2:{column_letter}{ws.max_row}',
            CellIsRule(
                operator='lessThanOrEqual',
                formula=[str(thresholds[0])],
                stopIfTrue=True,
                fill=PatternFill(start_color="AFF2C9", end_color="AFF2C9", fill_type="solid")
            )
        )

        # Yellow for needs improvement
        ws.conditional_formatting.add(
            f'{column_letter}2:{column_letter}{ws.max_row}',
            CellIsRule(
                operator='between',
                formula=[str(thresholds[0] + 0.0001), str(thresholds[1])],
                stopIfTrue=True,
                fill=PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
            )
        )

        # Red for poor performance
        ws.conditional_formatting.add(
            f'{column_letter}2:{column_letter}{ws.max_row}',
            CellIsRule(
                operator='greaterThan',
                formula=[str(thresholds[1])],
                stopIfTrue=True,
                fill=PatternFill(start_color="F7A9A9", end_color="F7A9A9", fill_type="solid")
            )
        )

    def _apply_borders(self, ws):
        """Apply borders to all data cells."""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border

    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
