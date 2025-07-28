"""
Excel Workbook Generator
Author: Karen Florykian
"""
import logging
from typing import List, Dict

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

from .core.data_models import (
    PerformanceReport,
    TransactionMetrics,
    PerformanceStatus,
    ThresholdConfig,
    ExcelFormattingConfig
)

logger = logging.getLogger(__name__)

class BusinessInsightsGenerator:
    """
    Generates actionable insights from performance data.
    Copilot-friendly: Simple methods for performance analysts to extend.
    """

    def __init__(self, config: ExcelFormattingConfig = None):
        """Initializes the generator with a formatting and thresholds configuration."""
        self.config = config or ExcelFormattingConfig()
        logger.info("BusinessInsightsGenerator initialized")

    def generate_key_insights(self, report: PerformanceReport) -> List[str]:
        """
        Generate key insights. Analysts can extend this by adding new private insight methods.
        """
        insights = [
            self._get_status_insight(report),
            self._get_error_rate_insight(report)
        ]
        insights.extend(self._get_throughput_insights(report))
        insights.extend(self._get_response_time_insights(report))

        # Filter out any None results from insight methods that might not apply
        final_insights = [insight for insight in insights if insight]
        logger.debug(f"Generated {len(final_insights)} insights for report")
        return final_insights

    def _get_status_insight(self, report: PerformanceReport) -> str:
        """Get overall status insight based on build status."""
        if report.build_status == PerformanceStatus.PASSED:
            return "✅ Overall system performance meets defined thresholds"
        return "🔴 System performance degradation detected - immediate attention required"

    def _get_throughput_insights(self, report: PerformanceReport) -> List[str]:
        """Get throughput-related insights."""
        insights = []
        # This threshold could also be moved to config if it varies
        if report.summary.throughput < 10.0:
            slow_transactions = [
                name for name, metrics in report.transactions.items()
                if name != "Total" and metrics.pct_95 > self.config.response_time_critical
            ]
            if slow_transactions:
                transaction_list = ", ".join(slow_transactions[:3])
                more_count = len(slow_transactions) - 3
                if more_count > 0:
                    transaction_list += f" and {more_count} more"
                insights.append(f"🔻 Throughput below 10.0 req/s affected by slow transactions: {transaction_list}")
        return insights

    def _get_error_rate_insight(self, report: PerformanceReport) -> str:
        """Get error rate insight."""
        if report.summary.error_rate > self.config.error_rate_warning:
            return f"⚠️ System error rate at {report.summary.error_rate:.2f}% - investigate failed transactions"
        return "✅ Zero error rate - all transactions completing successfully"

    def _get_response_time_insights(self, report: PerformanceReport) -> List[str]:
        """Get response time insights based on critical threshold."""
        insights = []
        critical_transactions = [
            name for name, metrics in report.transactions.items()
            if name != "Total" and metrics.pct_95 > self.config.response_time_critical
        ]
        if critical_transactions:
            insights.append(
                f"🔴 Response times > {self.config.response_time_critical / 1000}s by 95th percentile: "
                f"{len(critical_transactions)} transaction(s)"
            )
        return insights

    def filter_insights_by_intent(self, insights: List[str], query: str = "") -> List[str]:
        """
        PLACEHOLDER: Future intent recognition integration.
        For analysts: Add natural language processing here to filter insights.
        Example: "show only critical issues" -> filter insights with 🔴 emoji
        """
        if not query:
            return insights
        # Simple keyword-based filtering as a starting point
        logger.debug(f"Filtering insights with query: '{query}'")
        query_lower = query.lower()
        if "critical" in query_lower or "degradation" in query_lower:
            return [i for i in insights if "🔴" in i]
        if "warning" in query_lower or "investigate" in query_lower:
            return [i for i in insights if "⚠️" in i]
        return insights


# =================================================================================
# 2. EXCEL FORMATTER (Configuration-Driven)
# =================================================================================

class LegacyExcelFormatter:
    """
    Excel formatter using a configuration-driven approach.
    Copilot-friendly: Methods are organized by Excel feature for easy extension.
    """

    def __init__(self, config: ExcelFormattingConfig = None):
        """Initializes the formatter with a configuration object."""
        self.config = config or ExcelFormattingConfig()
        self._setup_fills_and_fonts()
        self.header_mapping = {
            'Transaction': 'request_name', 'Req, count': 'Total', 'KO, count': 'KO',
            'KO, %': 'Error_pct', 'Min, sec': 'min', 'Avg, sec': 'average',
            '90p, sec': 'pct_90', '95p, sec': 'pct_95', 'Max, sec': 'max', '📝 Notes': 'notes'
        }
        self.title_mapping = {
            'Users': 'max_user_count', 'Ramp Up, min': 'ramp_up_period', 'Duration, min': 'duration',
            'Think time, sec': 'think_time', 'Start Date, EST': 'date_start', 'End Date, EST': 'date_end',
            'Throughput, req/sec': 'throughput', 'Error rate, %': 'error_rate',
            'Carrier report': 'carrier_report', 'Hypothesis': 'hypothesis', 'Build status': 'build_status',
            'Justification': 'justification', 'Key Insights': 'key_insights',
            'Overall system performance': 'overall_performance'
        }
        logger.info("LegacyExcelFormatter initialized with configuration")

    def _setup_fills_and_fonts(self):
        """Creates openpyxl style objects from the configuration hex colors."""
        self.red_fill = PatternFill(
            start_color=self.config.critical_color.replace('#', ''),
            end_color=self.config.critical_color.replace('#', ''), fill_type='solid'
        )
        self.green_fill = PatternFill(
            start_color=self.config.good_color.replace('#', ''),
            end_color=self.config.good_color.replace('#', ''), fill_type='solid'
        )
        self.yellow_fill = PatternFill(
            start_color=self.config.warning_color.replace('#', ''),
            end_color=self.config.warning_color.replace('#', ''), fill_type='solid'
        )
        # Font colors for status indicators
        self.green_font_color = self.config.excellent_color.replace('#', '')
        self.red_font_color = self.config.critical_color_font.replace('#', '')

    def apply_title_section_formatting(self, ws: Worksheet, report: PerformanceReport, insights: List[str]):
        """Apply title section formatting using data from the report."""
        report_data = {
            'max_user_count': getattr(report.summary, 'max_user_count', 'N/A'),
            'ramp_up_period': getattr(report.summary, 'ramp_up_period', 'N/A'),
            'duration': getattr(report.summary, 'duration', 'N/A'),
            'think_time': getattr(report.summary, 'think_time', 'N/A'),
            'date_start': report.summary.date_start.strftime(
                '%Y-%m-%d %H:%M:%S') if report.summary.date_start else 'N/A',
            'date_end': report.summary.date_end.strftime('%Y-%m-%d %H:%M:%S') if report.summary.date_end else 'N/A',
            'throughput': report.summary.throughput,
            'error_rate': report.summary.error_rate / 100.0,
            'carrier_report': getattr(report.summary, 'carrier_report', 'N/A'),
            'hypothesis': getattr(report.summary, 'hypothesis', 'Validate against SLAs'),
            'build_status': report.build_status.value,
            'justification': report.analysis_summary or "Analysis not provided.",
            'key_insights': '\n'.join(insights),
            'overall_performance': 'HEALTHY' if report.build_status == PerformanceStatus.PASSED else 'DEGRADED'
        }

        border_style = Side(border_style="thin", color="040404")
        border = Border(top=border_style, left=border_style, right=border_style, bottom=border_style)
        title_fill = PatternFill("solid", fgColor='00CDEBEA')

        for i, (title_name, data_key) in enumerate(self.title_mapping.items(), 1):
            title_cell = ws.cell(row=i, column=1, value=title_name)
            title_cell.font = Font(bold=True, color='00291A75')
            title_cell.fill = title_fill
            title_cell.alignment = Alignment(horizontal="left", vertical="center")
            title_cell.border = border

            value_cell = ws.cell(row=i, column=2, value=report_data[data_key])
            value_cell.fill = title_fill
            value_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            value_cell.border = border

            if data_key == 'error_rate':
                value_cell.number_format = '0.00%'
            elif data_key == 'build_status':
                color = self.green_font_color if report_data[data_key] == 'PASSED' else self.red_font_color
                value_cell.font = Font(bold=True, color=color)
            elif data_key in ['justification', 'key_insights']:
                value_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                ws.row_dimensions[i].height = max(50, 15 * (str(report_data[data_key]).count('\n') + 1))

            ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=len(self.header_mapping))

        return len(self.title_mapping) + 1

    def apply_transaction_table_formatting(self, ws: Worksheet, report: PerformanceReport, start_row: int,
                                           thresholds: Dict[str, float]):
        """Apply transaction table formatting with config-driven conditional formatting."""
        header_row = start_row
        for col_idx, header_text in enumerate(self.header_mapping.keys(), 1):
            cell = ws.cell(row=header_row, column=col_idx, value=header_text)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill("solid", fgColor='007FD5D8')
            border_style = Side(border_style="thin", color="040404")
            cell.border = Border(top=border_style, left=border_style, right=border_style, bottom=border_style)

        current_row = header_row + 1
        sorted_transactions = sorted([name for name in report.transactions.keys() if name != "Total"])
        if "Total" in report.transactions:
            sorted_transactions.append("Total")

        for tx_name in sorted_transactions:
            metrics = report.transactions[tx_name]
            for col_idx, (header_name, data_key) in enumerate(self.header_mapping.items(), 1):
                cell_value = self._generate_validation_notes(metrics, thresholds) if data_key == 'notes' else getattr(
                    metrics, data_key, 'N/A')
                cell = ws.cell(row=current_row, column=col_idx, value=cell_value)
                border_style = Side(border_style="thin", color="040404")
                cell.border = Border(top=border_style, left=border_style, right=border_style, bottom=border_style)

                if data_key in ['min', 'average', 'pct_90', 'pct_95', 'max'] and isinstance(cell_value, (int, float)):
                    cell.value = round(cell_value / 1000, 3)
                elif data_key == 'Error_pct' and isinstance(cell_value, (int, float)):
                    cell.number_format = '0.00%'
                    cell.value = cell_value / 100.0

                if tx_name == "Total":
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor='007FD5D8')

            current_row += 1

        content_end_col = get_column_letter(len(self.header_mapping))
        self._apply_conditional_formatting(ws, header_row + 1, current_row - 1, thresholds)
        self._apply_legacy_table_features(ws, f"A{header_row}", f"{content_end_col}{current_row - 1}", header_row)
        return current_row

    def _generate_validation_notes(self, metrics: TransactionMetrics, thresholds: Dict[str, float]) -> str:
        """Generate validation notes based on response time thresholds from config."""
        logger.debug(f"Generating validation notes for {metrics.request_name}")
        rt_threshold = thresholds.get('response_time', self.config.response_time_warning)
        threshold_seconds = rt_threshold / 1000
        p95_seconds = metrics.pct_95 / 1000 if hasattr(metrics, 'pct_95') else 0

        if p95_seconds <= threshold_seconds:
            return "✅ Within SLA"
        else:
            pct_over = ((p95_seconds - threshold_seconds) / threshold_seconds) * 100
            return f"⚠️ {pct_over:.1f}% over SLA"

    def _apply_conditional_formatting(self, ws: Worksheet, start_row: int, end_row: int, thresholds: Dict[str, float]):
        """Apply conditional formatting using configuration-driven rules."""
        rt_threshold = thresholds.get('response_time', self.config.response_time_warning) / 1000  # Convert to seconds
        logger.debug(f"Applying conditional formatting with RT threshold: {rt_threshold}s")

        time_columns = [5, 6, 7, 8, 9]  # Columns E through I
        for col_idx in time_columns:
            col_letter = get_column_letter(col_idx)
            range_coords = f"{col_letter}{start_row}:{col_letter}{end_row}"

            ws.conditional_formatting.add(range_coords,
                                          CellIsRule(operator='lessThan', formula=[rt_threshold], fill=self.green_fill))
            ws.conditional_formatting.add(range_coords, CellIsRule(operator='greaterThan', formula=[rt_threshold * 1.5],
                                                                   fill=self.red_fill))
            ws.conditional_formatting.add(range_coords,
                                          CellIsRule(operator='between', formula=[rt_threshold, rt_threshold * 1.5],
                                                     fill=self.yellow_fill))

    def _apply_legacy_table_features(self, ws: Worksheet, content_start_ref: str, content_end_ref: str,
                                     header_row: int):
        """Apply legacy table features: auto-filter, freeze panes, and column widths."""
        logger.debug(f"Applying table features to range {content_start_ref}:{content_end_ref}")
        ws.auto_filter.ref = f"{content_start_ref}:{content_end_ref}"
        ws.freeze_panes = ws[f'B{header_row + 1}']
        for i, header in enumerate(self.header_mapping.keys(), 1):
            col_letter = get_column_letter(i)
            if header == "Transaction":
                max_length = max(len(str(cell.value)) for cell in ws[col_letter])
                ws.column_dimensions[col_letter].width = max(20, max_length + 5)
            else:
                ws.column_dimensions[col_letter].width = len(header) + 5


# =================================================================================
# 3. PRODUCTION EXCEL REPORTER (Main Orchestrator)
# =================================================================================

class ExcelReporter:
    """
    Production Excel reporter with configuration-driven formatting.
    Copilot-friendly: Clear separation of concerns for easy extension by analysts.
    """

    def __init__(self, config: ExcelFormattingConfig = None):
        """Initializes the reporter and its components with a shared configuration."""
        self.config = config or ExcelFormattingConfig()
        self.formatter = LegacyExcelFormatter(self.config)
        self.insights_generator = BusinessInsightsGenerator(self.config)
        logger.info("ExcelReporter initialized with configuration")

    def generate_workbook(self, report: PerformanceReport, analyst_query: str = "") -> Workbook:
        """
        Generate a production-ready Excel workbook.
        Optionally filters insights based on an analyst's natural language query.
        """
        logger.info(f"Generating Excel report for build status: {report.build_status.value}")
        wb = Workbook()
        ws = wb.active
        ws.title = "Test results"

        insights = self.insights_generator.generate_key_insights(report)
        if analyst_query:
            insights = self.insights_generator.filter_insights_by_intent(insights, analyst_query)

        thresholds = self._extract_thresholds(report.thresholds)
        next_row = self.formatter.apply_title_section_formatting(ws, report, insights)
        self.formatter.apply_transaction_table_formatting(ws, report, next_row, thresholds)

        logger.info("Production Excel report generated successfully.")
        return wb

    def generate_workbook_with_intent(self, report: PerformanceReport, query: str = "") -> Workbook:
        """
        PLACEHOLDER: Generate workbook filtered by analyst intent/query.
        This provides a clear entry point for future NLP-driven report customization.
        """
        logger.info(f"Generating intent-driven workbook with query: '{query}'")
        return self.generate_workbook(report, analyst_query=query)

    def _extract_thresholds(self, threshold_configs: List[ThresholdConfig]) -> Dict[str, float]:
        """Extract threshold values, using config defaults if not provided."""
        thresholds = {}
        if threshold_configs:
            for config in threshold_configs:
                thresholds[config.target] = config.threshold_value

        # Apply defaults from ExcelFormattingConfig if not explicitly set in the report
        thresholds.setdefault('response_time', self.config.response_time_warning)
        thresholds.setdefault('error_rate', self.config.error_rate_warning)

        logger.debug(f"Using thresholds for formatting: {thresholds}")
        return thresholds

    def _update_tests_overview_sheet(self, wb: Workbook, new_sheet_name: str, report: PerformanceReport):
        """Update or create a 'Tests Overview' sheet with a hyperlinked summary of the new report."""
        sheet_title = "Tests Overview"
        if sheet_title not in wb.sheetnames:
            ws = wb.create_sheet(sheet_title, 0)
            ws.append(["Test Date", "Report Sheet", "Build Status", "Throughput (req/s)", "P95 Response Time (ms)",
                       "Error Rate", "Key Issues"])
        else:
            ws = wb[sheet_title]

        key_issues = "None"
        if report.build_status != PerformanceStatus.PASSED:
            failed_insights = [i for i in self.insights_generator.generate_key_insights(report) if
                               "🔴" in i or "⚠️" in i]
            key_issues = ", ".join(i.split(' - ')[0] for i in failed_insights) or "Degradation detected"

        new_row = [
            report.summary.date_start.strftime('%Y-%m-%d %H:%M'),
            f'=HYPERLINK("#{new_sheet_name}!A1", "{new_sheet_name}")',
            report.build_status.value,
            report.summary.throughput,
            report.transactions.get("Total", TransactionMetrics(request_name="Total", Total=0, KO=0, OK=0, min=0, max=0,
                                                                average=0, median=0, Error_pct=0, pct_90=0,
                                                                pct_95=0)).pct_95,
            report.summary.error_rate / 100.0,
            key_issues
        ]
        ws.append(new_row)

        row_num = ws.max_row
        ws.cell(row=row_num, column=2).font = Font(bold=True, underline="single", color='00291A75')
        ws.cell(row=row_num, column=6).number_format = '0.00%'
        status_cell = ws.cell(row=row_num, column=3)
        if report.build_status == PerformanceStatus.PASSED:
            status_cell.font = Font(bold=True, color=self.formatter.green_font_color)
        else:
            status_cell.font = Font(bold=True, color=self.formatter.red_font_color)

    def update_report_with_new_sheet(self, workbook_path: str, new_report: PerformanceReport) -> bool:
        """Update an existing workbook with a new sheet containing the new report data."""
        try:
            wb = load_workbook(workbook_path)
            logger.info(f"Loaded existing workbook from '{workbook_path}'.")
        except FileNotFoundError:
            logger.warning(f"Workbook not found at '{workbook_path}'. Creating new workbook.")
            wb = Workbook()
            if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])

        sheet_name = new_report.summary.date_start.strftime('%Y-%m-%d_%H%M')
        if sheet_name in wb.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' already exists. Replacing it.")
            wb.remove(wb[sheet_name])

        ws_new = wb.create_sheet(sheet_name)

        insights = self.insights_generator.generate_key_insights(new_report)
        thresholds = self._extract_thresholds(new_report.thresholds)

        next_row = self.formatter.apply_title_section_formatting(ws_new, new_report, insights)
        self.formatter.apply_transaction_table_formatting(ws_new, new_report, next_row, thresholds)

        self._update_tests_overview_sheet(wb, sheet_name, new_report)

        wb.save(workbook_path)
        logger.info(f"Successfully updated workbook saved to '{workbook_path}'.")
        return True
